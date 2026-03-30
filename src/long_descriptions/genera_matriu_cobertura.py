from __future__ import annotations

import math
import random
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_XLSX = Path("matrius_cobertura_excel_3350_335.xlsx")
OUTPUT_CSV_5000 = Path("matriu_3350.csv")
OUTPUT_CSV_500 = Path("matriu_335.csv")
RANDOM_SEED = 20260311


FAMILIES: List[Dict] = [
    {
        "code": "C01",
        "family": "Column",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C01.01", "Clustered Column", 360, False),
            ("C01.02", "Stacked Column", 180, False),
            ("C01.03", "100% Stacked Column", 120, False),
        ],
        "structures": ["categòric simple", "categòric multiserie", "sèrie temporal discreta"],
        "tasks": ["llegir valor", "comparar categories", "trobar màxim/mínim", "ordenar/ranquejar", "estimar diferència"],
        "patterns": ["diferències clares", "valors molt propers", "pic local", "creixement", "decreixement", "valors negatius"],
        "domains": ["vendes", "finances", "educació", "salut", "energia", "operacions", "manufactura"],
        "styles": ["estàndard", "llegenda a la dreta", "etiquetes de dades", "quadrícula lleu"],
    },
    {
        "code": "C02",
        "family": "Bar",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C02.01", "Clustered Bar", 150, False),
            ("C02.02", "Stacked Bar", 90, False),
            ("C02.03", "100% Stacked Bar", 60, False),
        ],
        "structures": ["categòric simple", "categòric multiserie", "categòric amb etiquetes llargues"],
        "tasks": ["llegir valor", "comparar categories", "trobar màxim/mínim", "ordenar/ranquejar", "estimar diferència"],
        "patterns": ["diferències clares", "valors molt propers", "long tail", "pic local", "valors negatius"],
        "domains": ["vendes", "salut", "educació", "demografia", "operacions", "web analytics"],
        "styles": ["estàndard", "etiquetes llargues", "llegenda a baix", "etiquetes de dades"],
    },
    {
        "code": "C03",
        "family": "Line",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C03.01", "Line", 220, False),
            ("C03.02", "Line with Markers", 170, False),
            ("C03.03", "Stacked Line", 110, False),
            ("C03.04", "Stacked Line with Markers", 100, False),
            ("C03.05", "100% Stacked Line", 90, False),
        ],
        "structures": ["sèrie temporal", "multiserie temporal", "categòric ordenat"],
        "tasks": ["llegir valor", "comparar sèries", "trobar màxim/mínim", "detectar tendència", "detectar canvi de règim", "detectar estacionalitat"],
        "patterns": ["creixement", "decreixement", "estacionalitat", "canvi de règim", "soroll alt", "pic sobtat", "vall sobtada"],
        "domains": ["vendes", "finances", "energia", "clima", "web analytics", "operacions"],
        "styles": ["estàndard", "amb marcadors", "sense quadrícula", "llegenda a la dreta"],
    },
    {
        "code": "C04",
        "family": "Pie",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C04.01", "Pie", 80, False),
            ("C04.02", "Pie of Pie", 40, False),
            ("C04.03", "Bar of Pie", 30, False),
            ("C04.04", "Exploded Pie", 30, False),
        ],
        "structures": ["parts del tot simple", "parts del tot amb cua llarga"],
        "tasks": ["comparar proporcions", "identificar segment dominant", "estimar quota", "detectar long tail"],
        "patterns": ["parts equilibrades", "part dominant", "cua llarga", "segments molt propers"],
        "domains": ["vendes", "web analytics", "demografia", "educació", "operacions"],
        "styles": ["estàndard", "amb percentatges", "llegenda a la dreta", "segment ressaltat"],
    },
    {
        "code": "C05",
        "family": "Doughnut",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [("C05.01", "Doughnut", 110, False)],
        "structures": ["parts del tot simple", "parts del tot multianell"],
        "tasks": ["comparar proporcions", "identificar segment dominant", "comparar anells"],
        "patterns": ["parts equilibrades", "part dominant", "segments molt propers"],
        "domains": ["vendes", "web analytics", "operacions", "demografia"],
        "styles": ["estàndard", "amb percentatges", "anell doble", "segment ressaltat"],
    },
    {
        "code": "C06",
        "family": "Area",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C06.01", "Area", 70, False),
            ("C06.02", "Stacked Area", 50, False),
            ("C06.03", "100% Stacked Area", 30, False),
        ],
        "structures": ["sèrie temporal", "multiserie temporal acumulada"],
        "tasks": ["detectar tendència", "comparar magnituds", "analitzar acumulació", "comparar composició al llarg del temps"],
        "patterns": ["creixement", "decreixement", "estacionalitat", "canvi de règim", "acumulació creixent"],
        "domains": ["energia", "clima", "web analytics", "operacions", "vendes"],
        "styles": ["estàndard", "amb transparència", "llegenda a la dreta", "sense quadrícula"],
    },
    {
        "code": "C07",
        "family": "Scatter",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C07.01", "Scatter with Only Markers", 120, False),
            ("C07.02", "Scatter with Smooth Lines", 70, False),
            ("C07.03", "Scatter with Smooth Lines and Markers", 90, False),
            ("C07.04", "Scatter with Straight Lines", 60, False),
        ],
        "structures": ["bivariant numèrica"],
        "tasks": ["detectar correlació", "detectar outlier", "detectar clústers", "comparar punts", "detectar relació no lineal"],
        "patterns": ["correlació positiva", "correlació negativa", "correlació nul·la", "outlier clar", "clústers", "relació corba"],
        "domains": ["salut", "energia", "clima", "demografia", "manufactura", "educació"],
        "styles": ["estàndard", "amb línia de tendència", "marcadors grans", "quadrícula lleu"],
    },
    {
        "code": "C09",
        "family": "Stock",
        "block": "Especialitzat",
        "modern": False,
        "subtypes": [
            ("C09.01", "High-Low-Close", 50, False),
            ("C09.02", "Open-High-Low-Close", 50, False),
            ("C09.03", "Volume-High-Low-Close", 40, False),
            ("C09.04", "Volume-Open-High-Low-Close", 50, False),
        ],
        "structures": ["OHLC temporal", "OHLC + volum"],
        "tasks": ["trobar màxim/mínim", "comparar obertura i tancament", "detectar volatilitat", "detectar tendència"],
        "patterns": ["alta volatilitat", "baixa volatilitat", "tendència alcista", "tendència baixista", "gap puntual"],
        "domains": ["finances", "energia"],
        "styles": ["estàndard", "volum visible", "sense quadrícula", "llegenda a la dreta"],
    },
    {
        "code": "C10",
        "family": "Surface",
        "block": "Estructurat",
        "modern": False,
        "subtypes": [
            ("C10.01", "Contour", 10, False),
            ("C10.02", "Wireframe Contour", 10, False),
        ],
        "structures": ["graella numèrica 2D"],
        "tasks": ["trobar pic màxim", "trobar vall mínima", "detectar gradient", "comparar regions", "identificar crestes"],
        "patterns": ["superfície suau", "superfície rugosa", "pic central", "doble pic", "gradient diagonal"],
        "domains": ["clima", "energia", "manufactura"],
        "styles": ["estàndard", "mallat visible", "línies de contorn", "paleta accent"],
    },
    {
        "code": "C11",
        "family": "Radar",
        "block": "Estructurat",
        "modern": False,
        "subtypes": [("C11.01", "Radar", 30, False), ("C11.02", "Radar with Markers", 30, False), ("C11.03", "Filled Radar", 20, False)],
        "structures": ["multivariable sobre eixos comuns"],
        "tasks": ["comparar perfils", "identificar dimensió dominant", "comparar sèries", "detectar perfil equilibrat/desequilibrat"],
        "patterns": ["perfil equilibrat", "perfil espigat", "dues sèries contrastades", "una dimensió dominant"],
        "domains": ["educació", "salut", "operacions", "manufactura"],
        "styles": ["estàndard", "amb marcadors", "omplert", "llegenda a baix"],
    },
    {
        "code": "C14",
        "family": "Histogram",
        "block": "Especialitzat",
        "modern": True,
        "subtypes": [("C14.01", "Histogram", 250, False)],
        "structures": ["distribució univariant"],
        "tasks": ["caracteritzar distribució", "detectar asimetria", "detectar concentració", "detectar outlier", "comparar dispersió"],
        "patterns": ["simètrica", "asimètrica", "bimodal", "cua llarga", "outlier clar"],
        "domains": ["salut", "educació", "demografia", "energia", "manufactura"],
        "styles": ["estàndard", "més bins", "menys bins", "sense quadrícula"],
    },
    {
        "code": "C15",
        "family": "Pareto",
        "block": "Especialitzat",
        "modern": True,
        "subtypes": [("C15.01", "Pareto", 100, False)],
        "structures": ["categòric ordenat + acumulat"],
        "tasks": ["identificar categories dominants", "trobar llindar 80/20", "ordenar/ranquejar", "interpretar acumulat"],
        "patterns": ["80/20 marcat", "cua llarga", "concentració moderada", "concentració forta"],
        "domains": ["operacions", "manufactura", "vendes", "web analytics"],
        "styles": ["estàndard", "línia acumulada destacada", "etiquetes de dades", "sense quadrícula"],
    },
    {
        "code": "C18",
        "family": "Funnel",
        "block": "Especialitzat",
        "modern": True,
        "subtypes": [("C18.01", "Funnel", 100, False)],
        "structures": ["procés per etapes"],
        "tasks": ["identificar bottleneck", "calcular caiguda més gran", "comparar etapes", "interpretar conversió"],
        "patterns": ["caiguda uniforme", "bottleneck clar", "caiguda tardana", "caiguda inicial forta"],
        "domains": ["web analytics", "vendes", "operacions", "salut"],
        "styles": ["estàndard", "percentatges visibles", "etapes etiquetades", "paleta accent"],
    },
    {
        "code": "C19",
        "family": "Combo",
        "block": "Nuclear",
        "modern": True,
        "subtypes": [
            ("C19.01", "Clustered Column + Line", 80, False),
            ("C19.02", "Column + Line on Secondary Axis", 60, False),
            ("C19.03", "Custom Combo", 40, False),
        ],
        "structures": ["multiserie mixta", "temporal amb eix secundari"],
        "tasks": ["comparar sèries", "detectar relació entre volum i taxa", "detectar tendència", "interpretar eix secundari"],
        "patterns": ["creixement amb taxa", "volum i percentatge", "dues escales diferents", "sèrie principal + objectiu"],
        "domains": ["vendes", "finances", "web analytics", "energia", "operacions"],
        "styles": ["estàndard", "eix secundari visible", "llegenda a la dreta", "etiquetes selectives"],
    },
]


def difficulty_sequence(count: int, block: str) -> List[str]:
    if block == "Nuclear":
        ratios = [0.35, 0.40, 0.20, 0.05]
    elif block == "Especialitzat":
        ratios = [0.20, 0.35, 0.30, 0.15]
    else:
        ratios = [0.15, 0.35, 0.35, 0.15]

    labels = ["baixa", "mitjana", "alta", "molt alta"]
    raw = [count * r for r in ratios]
    base = [math.floor(x) for x in raw]

    while sum(base) < count:
        remainders = [raw[i] - base[i] for i in range(4)]
        idx = max(range(4), key=lambda i: (remainders[i], -i))
        base[idx] += 1

    seq: List[str] = []
    pools = {lab: n for lab, n in zip(labels, base)}
    order = ["mitjana", "baixa", "alta", "mitjana", "baixa", "alta", "molt alta"]

    while len(seq) < count:
        for lab in order:
            if pools[lab] > 0 and len(seq) < count:
                seq.append(lab)
                pools[lab] -= 1
    return seq


def question_template(task: str) -> Tuple[str, str]:
    mapping = {
        "llegir valor": ("Quin valor té {element_objetiu} al gràfic?", "valor_numèric"),
        "comparar categories": ("Quina categoria és més gran entre {element_A} i {element_B}?", "comparació_categòrica"),
        "comparar sèries": ("Quina sèrie és superior al punt/categoria {element_objetiu}?", "comparació_sèries"),
        "trobar màxim/mínim": ("Quin element mostra el valor màxim o mínim?", "element_extrem"),
        "ordenar/ranquejar": ("Quins són els elements principals en ordre descendent?", "rànquing"),
        "estimar diferència": ("Quina diferència aproximada hi ha entre {element_A} i {element_B}?", "diferència"),
        "detectar tendència": ("La tendència general és creixent, decreixent o estable?", "patró_tendència"),
        "detectar canvi de règim": ("En quin punt canvia clarament el comportament de la sèrie?", "punt_canvi"),
        "detectar estacionalitat": ("S'observa un patró estacional o cíclic?", "patró_estacional"),
        "comparar proporcions": ("Quin segment representa una proporció més gran del total?", "segment_dominant"),
        "identificar segment dominant": ("Quin segment domina el total?", "segment_dominant"),
        "estimar quota": ("Quina quota aproximada representa {element_objetiu}?", "percentatge"),
        "detectar long tail": ("Hi ha una cua llarga de segments petits?", "patró_composició"),
        "comparar magnituds": ("Quina sèrie o interval té la magnitud més alta?", "comparació_magnitud"),
        "analitzar acumulació": ("Com evoluciona el total acumulat al llarg de la sèrie?", "acumulació"),
        "comparar composició al llarg del temps": ("Com canvia la composició relativa al llarg del temps?", "composició_temporal"),
        "detectar correlació": ("Hi ha correlació positiva, negativa o nul·la?", "correlació"),
        "detectar outlier": ("Hi ha algun punt o valor clarament anòmal?", "outlier"),
        "detectar outliers": ("Hi ha algun valor atípic o outlier visible?", "outlier"),
        "detectar clústers": ("S'observen agrupacions o clústers diferenciats?", "clústers"),
        "comparar punts": ("Quin punt o grup de punts destaca més?", "comparació_punts"),
        "detectar relació no lineal": ("La relació sembla lineal o corba?", "forma_relació"),
        "comparar magnitud de bombolles": ("Quina bombolla representa la mida més gran?", "mida_relativa"),
        "comparar obertura i tancament": ("En quin període la diferència entre obertura i tancament és més gran?", "diferència_obertura_tancament"),
        "detectar volatilitat": ("En quin tram s'observa més volatilitat?", "volatilitat"),
        "trobar pic màxim": ("On es troba el pic màxim de la superfície?", "regió_extrema"),
        "trobar vall mínima": ("On es troba la vall mínima de la superfície?", "regió_extrema"),
        "detectar gradient": ("Quina direcció mostra el gradient principal?", "gradient"),
        "comparar regions": ("Quina regió o zona mostra el valor més alt?", "regió_extrema"),
        "identificar crestes": ("S'identifiquen crestes o zones elevades diferenciades?", "morfologia_superfície"),
        "comparar perfils": ("Quin perfil és més equilibrat o més extrem?", "comparació_perfils"),
        "identificar dimensió dominant": ("Quina dimensió o eix domina el perfil?", "dimensió_dominant"),
        "detectar perfil equilibrat/desequilibrat": ("El perfil és equilibrat o està dominat per pocs eixos?", "perfil_global"),
        "identificar branca dominant": ("Quina branca jeràrquica acumula més pes?", "branca_dominant"),
        "comparar parts del tot": ("Quin subgrup representa una part més gran del total?", "comparació_parts"),
        "detectar jerarquia": ("Quina és l'estructura jeràrquica principal del gràfic?", "jerarquia"),
        "comparar subgrups": ("Quin subgrup és més gran dins de la branca principal?", "comparació_subgrups"),
        "comparar anells jeràrquics": ("Quin nivell jeràrquic concentra més valor?", "nivell_dominant"),
        "caracteritzar distribució": ("La distribució és concentrada, dispersa, simètrica o asimètrica?", "descripció_distribució"),
        "detectar asimetria": ("La distribució és simètrica o té cua cap a un costat?", "asimetria"),
        "detectar concentració": ("La major part dels valors es concentra en un rang estret?", "concentració"),
        "comparar dispersió": ("Quin grup o distribució mostra més dispersió?", "dispersió"),
        "identificar categories dominants": ("Quines categories concentren la major part de l'impacte?", "categories_dominants"),
        "trobar llindar 80/20": ("Quantes categories calen per arribar aproximadament al 80% acumulat?", "llindar_80_20"),
        "interpretar acumulat": ("Com evoluciona la línia acumulada o el total acumulat?", "acumulació"),
        "comparar medianes": ("Quin grup té la mediana més alta?", "mediana"),
        "comparar asimetria": ("Quin grup mostra més asimetria?", "asimetria"),
        "identificar contribució principal": ("Quin pas contribueix més al resultat final?", "pas_dominant"),
        "trobar pas més positiu/negatiu": ("Quin pas té l'efecte positiu o negatiu més gran?", "pas_extrem"),
        "comparar saldo inicial i final": ("Com canvia el total entre l'inici i el final?", "canvi_total"),
        "identificar bottleneck": ("En quina etapa es produeix la major pèrdua?", "bottleneck"),
        "calcular caiguda més gran": ("Entre quines etapes hi ha la caiguda més gran?", "caiguda_màxima"),
        "comparar etapes": ("Quina etapa reté més volum relatiu?", "comparació_etapes"),
        "interpretar conversió": ("Quina és la conversió aproximada fins a l'etapa final?", "conversió"),
        "detectar relació entre volum i taxa": ("La línia i les columnes mostren un comportament alineat o divergent?", "relació_sèries"),
        "interpretar eix secundari": ("Quina sèrie s'està interpretant sobre l'eix secundari i què implica?", "eix_secundari"),
        "detectar hotspot": ("Hi ha alguna regió que actuï com a hotspot clar?", "hotspot"),
        "detectar gradient espacial": ("S'observa un gradient espacial recognoscible?", "gradient_espacial"),
        "identificar regió extrema": ("Quina regió mostra l'extrem principal?", "regió_extrema"),
    }
    return mapping.get(task, ("Quina és la lectura principal que s'obté d'aquest gràfic?", "resposta_qualitativa"))


def style_sequence(count: int, styles: List[str], family: str | None = None) -> List[str]:
    base = list(styles)
    return [base[i % len(base)] for i in range(count)]


def family_summary(families: List[Dict]) -> List[Dict]:
    rows = []
    for fam in families:
        q5000 = sum(st[2] for st in fam["subtypes"])
        rows.append(
            {
                "codi_familia": fam["code"],
                "familia_excel": fam["family"],
                "bloc_cobertura": fam["block"],
                "n_subtipus": len(fam["subtypes"]),
                "es_modern": "Yes" if fam["modern"] else "No",
                "quota_5000": q5000,
                "quota_500": q5000 // 10,
                "pes_5000": q5000 / 5000,
                "pes_500": (q5000 // 10) / 500,
            }
        )
    return rows


def generate_cases(families: List[Dict], seed: int = RANDOM_SEED, target_full: int = 3350, target_sample: int = 335) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    rng = random.Random(seed)
    full_cases: List[Dict] = []
    coverage_rows: List[Dict] = []
    case_num = 1
    
    # Calculate base total from family quotas
    base_total = sum(q for fam in families for _, _, q, _ in fam["subtypes"])
    scale_factor = target_full / base_total if base_total > 0 else 1.0

    # First pass: calculate scaled quotas
    quota_map = {}
    for fam in families:
        for subtype_code, subtype, quota_base, is_3d in fam["subtypes"]:
            quota_full = max(1, round(quota_base * scale_factor))
            quota_sample = max(1, round(quota_full * target_sample / target_full))
            quota_map[(fam["code"], subtype_code)] = {
                "quota_full": quota_full,
                "quota_sample": quota_sample,
                "quota_base": quota_base
            }
    
    # Adjust to hit exact target
    current_total = sum(q["quota_full"] for q in quota_map.values())
    diff = target_full - current_total
    
    if diff != 0:
        # Sort by base quota (adjust largest items first)
        sorted_keys = sorted(quota_map.keys(), key=lambda k: quota_map[k]["quota_base"], reverse=(diff > 0))
        for i, key in enumerate(sorted_keys):
            if diff == 0:
                break
            adjustment = 1 if diff > 0 else -1
            if quota_map[key]["quota_full"] + adjustment >= 1:
                quota_map[key]["quota_full"] += adjustment
                # Proportionally adjust sample quota
                new_sample = max(1, round(quota_map[key]["quota_full"] * target_sample / target_full))
                quota_map[key]["quota_sample"] = new_sample
                diff -= adjustment

    # Second pass: ensure sample total is correct
    current_sample_total = sum(q["quota_sample"] for q in quota_map.values())
    sample_diff = target_sample - current_sample_total
    
    if sample_diff != 0:
        sorted_keys = sorted(quota_map.keys(), key=lambda k: quota_map[k]["quota_base"], reverse=(sample_diff > 0))
        for key in sorted_keys:
            if sample_diff == 0:
                break
            adjustment = 1 if sample_diff > 0 else -1
            if quota_map[key]["quota_sample"] + adjustment >= 1:
                quota_map[key]["quota_sample"] += adjustment
                sample_diff -= adjustment

    for fam in families:
        for subtype_code, subtype, quota_base, is_3d in fam["subtypes"]:
            quotas = quota_map[(fam["code"], subtype_code)]
            quota_full = quotas["quota_full"]
            quota_sample = quotas["quota_sample"]
            difficulties = difficulty_sequence(quota_full, fam["block"])
            styles = style_sequence(quota_full, fam["styles"], fam["family"])

            structures = [fam["structures"][(idx * 2 + 1) % len(fam["structures"])] for idx in range(quota_full)]
            tasks = [fam["tasks"][(idx * 3 + 2) % len(fam["tasks"])] for idx in range(quota_full)]
            patterns = [fam["patterns"][(idx * 5 + 1) % len(fam["patterns"])] for idx in range(quota_full)]
            domains = [fam["domains"][(idx * 7 + 3) % len(fam["domains"])] for idx in range(quota_full)]

            if subtype in {"Pie of Pie", "Bar of Pie"}:
                structures = ["parts del tot amb cua llarga"] * quota_full
                patterns = ["cua llarga"] * quota_full
            if fam["family"] == "Surface":
                structures = ["graella numèrica 2D"] * quota_full
            if fam["family"] == "Histogram":
                structures = ["distribució univariant"] * quota_full

            selected_orders = set(sorted(rng.sample(range(1, quota_full + 1), quota_sample)))

            coverage_rows.append(
                {
                    "codi_familia": fam["code"],
                    "familia_excel": fam["family"],
                    "codi_subtipus": subtype_code,
                    "subtipus_excel": subtype,
                    "bloc_cobertura": fam["block"],
                    "es_modern": "Yes" if fam["modern"] else "No",
                    "es_3d": "Yes" if is_3d else "No",
                    "estructura_principal": ", ".join(fam["structures"]),
                    "tasques_compatibles": ", ".join(fam["tasks"]),
                    "patrons_compatibles": ", ".join(fam["patterns"]),
                    "quota_5000": quota_full,
                    "quota_500": quota_sample,
                    "pes_5000": quota_full / target_full,
                    "pes_500": quota_sample / target_sample,
                }
            )

            for idx in range(quota_full):
                task = tasks[idx]
                prompt, answer_type = question_template(task)
                order = idx + 1
                full_cases.append(
                    {
                        "case_id": f"CASE_{case_num:05d}",
                        "codi_familia": fam["code"],
                        "familia_excel": fam["family"],
                        "codi_subtipus": subtype_code,
                        "subtipus_excel": subtype,
                        "bloc_cobertura": fam["block"],
                        "estructura_dades": structures[idx],
                        "patro_estadistic": patterns[idx],
                        "tasca_analitica": task,
                        "plantilla_pregunta": prompt,
                        "tipus_resposta_oracle": answer_type,
                        "dificultat": difficulties[idx],
                        "domini_semantic": domains[idx],
                        "variant_estil": styles[idx],
                        "es_modern": "Yes" if fam["modern"] else "No",
                        "quota_subtipus_5000": quota_full,
                        "quota_subtipus_500": quota_sample,
                        "ordre_dins_subtipus": order,
                        "seleccionat_a_mostra_500": "Yes" if order in selected_orders else "No",
                    }
                )
                case_num += 1

    sample_cases = [row for row in full_cases if row["seleccionat_a_mostra_500"] == "Yes"]
    return full_cases, sample_cases, coverage_rows


# ---------- Excel helpers ----------
fill_title = PatternFill("solid", fgColor="1F4E78")
fill_subheader = PatternFill("solid", fgColor="EAF3F8")
fill_nuclear = PatternFill("solid", fgColor="DDEBF7")
fill_special = PatternFill("solid", fgColor="FFF2CC")
fill_struct = PatternFill("solid", fgColor="E4DFEC")
fill_modern = PatternFill("solid", fgColor="E2F0D9")
fill_3d = PatternFill("solid", fgColor="FCE4D6")
white_font = Font(color="FFFFFF", bold=True, size=12)
title_font = Font(color="FFFFFF", bold=True, size=14)
header_font = Font(bold=True, color="1F1F1F")
normal_font = Font(color="000000", size=10)
blue_font = Font(color="1F4E78", bold=True)
border_light = Border(bottom=Side(style="thin", color="D9D9D9"))


def add_sheet(wb: Workbook, name: str, headers: List[str], rows: List[Dict], widths: Dict[int, int] | None = None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill_title
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_light

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = normal_font
            cell.border = border_light

    if widths:
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    return ws


def build_workbook(full_cases: List[Dict], sample_cases: List[Dict], coverage_rows: List[Dict], output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    fam_summary = family_summary(FAMILIES)

    # Overview sheet
    ws = wb.create_sheet("Overview")
    ws["A1"] = "Coverage Matrix for Excel Charts"
    ws["A1"].fill = fill_title
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")

    lines = [
        ("Objective", "Define comprehensive coverage of Excel chart types with a complete matrix of 5,000 cases and an initial proportional sample of 500."),
        ("Sampling criteria", "The sample of 500 is exactly 10% of each subtype from the 5,000 matrix."),
        ("Selection method", "Stratified and deterministic random selection within each subtype (fixed seed)."),
        ("Coverage level", "Coverage is organized by family, subtype, data structure, pattern, task, difficulty, and semantic domain."),
        ("Subtypes", "60 operational subtypes have been modeled."),
    ]
    row = 3
    for label, text in lines:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = blue_font
        ws[f"A{row}"].fill = fill_subheader
        ws[f"B{row}"] = text
        row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    ws.sheet_view.showGridLines = False

    # Family summary
    fam_headers = [
        "codi_familia", "familia_excel", "bloc_cobertura", "n_subtipus", "es_modern",
        "quota_objectiu_5000", "actual_5000", "quota_objectiu_500", "actual_500",
        "pes_5000", "pes_500",
    ]
    wsf = wb.create_sheet("Family_Summary")
    wsf.append(fam_headers)
    for c in wsf[1]:
        c.fill = fill_title
        c.font = white_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border_light

    for i, fam in enumerate(fam_summary, start=2):
        wsf[f"A{i}"] = fam["codi_familia"]
        wsf[f"B{i}"] = fam["familia_excel"]
        wsf[f"C{i}"] = fam["bloc_cobertura"]
        wsf[f"D{i}"] = fam["n_subtipus"]
        wsf[f"E{i}"] = fam["es_modern"]
        wsf[f"F{i}"] = fam["quota_5000"]
        wsf[f"G{i}"] = f'=COUNTIF(Matrix_3350!B:B,A{i})'
        wsf[f"H{i}"] = fam["quota_500"]
        wsf[f"I{i}"] = f'=COUNTIF(Matrix_335!B:B,A{i})'
        wsf[f"J{i}"] = f'=G{i}/5000'
        wsf[f"K{i}"] = f'=I{i}/500'

    wsf.freeze_panes = "A2"
    wsf.auto_filter.ref = f"A1:K{wsf.max_row}"
    wsf.sheet_view.showGridLines = False

    # Subtype summary
    sub_headers = [
        "codi_familia", "familia_excel", "codi_subtipus", "subtipus_excel", "bloc_cobertura", "es_modern",         "estructura_principal", "tasques_compatibles", "patrons_compatibles",
        "quota_objectiu_5000", "actual_5000", "quota_objectiu_500", "actual_500", "pes_5000", "pes_500",
    ]
    wss = wb.create_sheet("Subtype_Summary")
    wss.append(sub_headers)
    for c in wss[1]:
        c.fill = fill_title
        c.font = white_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border_light

    for i, rowd in enumerate(coverage_rows, start=2):
        wss[f"A{i}"] = rowd["codi_familia"]
        wss[f"B{i}"] = rowd["familia_excel"]
        wss[f"C{i}"] = rowd["codi_subtipus"]
        wss[f"D{i}"] = rowd["subtipus_excel"]
        wss[f"E{i}"] = rowd["bloc_cobertura"]
        wss[f"F{i}"] = rowd["es_modern"]
        wss[f"H{i}"] = rowd["estructura_principal"]
        wss[f"I{i}"] = rowd["tasques_compatibles"]
        wss[f"J{i}"] = rowd["patrons_compatibles"]
        wss[f"K{i}"] = rowd["quota_5000"]
        wss[f"L{i}"] = f'=COUNTIF(Matrix_3350!D:D,C{i})'
        wss[f"M{i}"] = rowd["quota_500"]
        wss[f"N{i}"] = f'=COUNTIF(Matrix_335!D:D,C{i})'
        wss[f"O{i}"] = f'=L{i}/5000'
        wss[f"P{i}"] = f'=N{i}/500'

    wss.freeze_panes = "A2"
    wss.auto_filter.ref = f"A1:P{wss.max_row}"
    wss.sheet_view.showGridLines = False

    # Matrices
    matrix_headers = list(full_cases[0].keys())
    add_sheet(
        wb,
        "Matrix_3350",
        matrix_headers,
        full_cases,
        widths={1: 12, 2: 10, 3: 14, 4: 12, 5: 28, 6: 16, 7: 18, 8: 22, 9: 28, 10: 48, 11: 20, 12: 12, 13: 16, 14: 18, 15: 10, 16: 8, 17: 16, 18: 16, 19: 16, 20: 14},
    )
    add_sheet(
        wb,
        "Matrix_335",
        matrix_headers,
        sample_cases,
        widths={1: 12, 2: 10, 3: 14, 4: 12, 5: 28, 6: 16, 7: 18, 8: 22, 9: 28, 10: 48, 11: 20, 12: 12, 13: 16, 14: 18, 15: 10, 16: 8, 17: 16, 18: 16, 19: 16, 20: 14},
    )

    # Colors per block and flags
    for sheet_name in ["Family_Summary", "Subtype_Summary", "Matrix_3350", "Matrix_335"]:
        wsx = wb[sheet_name]
        start = 2
        end = wsx.max_row
        for r in range(start, end + 1):
            block_col = 3 if sheet_name == "Family_Summary" else 5 if sheet_name == "Subtype_Summary" else 6
            modern_col = 5 if sheet_name == "Family_Summary" else 6 if sheet_name == "Subtype_Summary" else 15
            d3_col = None if sheet_name == "Family_Summary" else 7 if sheet_name == "Subtype_Summary" else 16
            block = wsx.cell(r, block_col).value
            fill = fill_nuclear if block == "Nuclear" else fill_special if block == "Especialitzat" else fill_struct
            if sheet_name == "Family_Summary":
                cols = [1, 2, 3]
            elif sheet_name == "Subtype_Summary":
                cols = [1, 2, 3, 4]
            else:
                cols = [2, 3, 4, 5, 6]
            for c in cols:
                wsx.cell(r, c).fill = fill
            if wsx.cell(r, modern_col).value == "Yes":
                wsx.cell(r, modern_col).fill = fill_modern
            if d3_col and wsx.cell(r, d3_col).value == "Yes":
                wsx.cell(r, d3_col).fill = fill_3d

    wb.save(output_path)

### ---------- Main execution ----------

###
def main(target_full_cases: int = 5000, target_sample_cases: int = 500) -> None:
    full_cases, sample_cases, coverage_rows = generate_cases(FAMILIES, seed=RANDOM_SEED, target_full=target_full_cases, target_sample=target_sample_cases)

    assert len(full_cases) == target_full_cases, f"Expected {target_full_cases} cases but got {len(full_cases)}"
    assert len(sample_cases) == target_sample_cases, f"Expected {target_sample_cases} cases but got {len(sample_cases)}"

    output_xlsx = Path(f"matrius_cobertura_excel_{target_full_cases}_{target_sample_cases}.xlsx")
    output_csv_full = Path(f"matriu_{target_full_cases}.csv")
    output_csv_sample = Path(f"matriu_{target_sample_cases}.csv")

    build_workbook(full_cases, sample_cases, coverage_rows, output_xlsx)
    pd.DataFrame(full_cases).to_csv(output_csv_full, index=False, encoding="utf-8-sig")
    pd.DataFrame(sample_cases).to_csv(output_csv_sample, index=False, encoding="utf-8-sig")

    print(f"Excel file created: {output_xlsx.resolve()}")
    print(f"CSV {target_full_cases} created:   {output_csv_full.resolve()}")
    print(f"CSV {target_sample_cases} created:    {output_csv_sample.resolve()}")


if __name__ == "__main__":
    main()
