from __future__ import annotations

import math
import random
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("output")
OUTPUT_XLSX = OUTPUT_DIR / "coverage_matrix_5000_500.xlsx"
OUTPUT_CSV_5000 = OUTPUT_DIR / "matrix_5000.csv"
OUTPUT_CSV_500 = OUTPUT_DIR / "matrix_500.csv"
RANDOM_SEED = 20260311


FAMILIES: List[Dict] = [
    {
        "code": "C01",
        "family": "Column",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C01.01", "Clustered Column", 360, False),
            ("C01.02", "Stacked Column", 180, False),
            ("C01.03", "100% Stacked Column", 120, False),
        ],
        "structures": ["simple categorical", "multi-series categorical", "discrete time series"],
        "tasks": ["read value", "compare categories", "find max/min", "order/rank", "estimate difference"],
        "patterns": ["clear differences", "very close values", "local peak", "growth", "decline", "negative values"],
        "domains": ["sales", "finance", "education", "health", "energy", "operations", "manufacturing"],
        "styles": ["standard", "legend on the right", "data labels", "light grid"],
    },
    {
        "code": "C02",
        "family": "Bar",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C02.01", "Clustered Bar", 150, False),
            ("C02.02", "Stacked Bar", 90, False),
            ("C02.03", "100% Stacked Bar", 60, False),
        ],
        "structures": ["simple categorical", "multi-series categorical", "categorical with long labels"],
        "tasks": ["read value", "compare categories", "find max/min", "order/rank", "estimate difference"],
        "patterns": ["clear differences", "very close values", "long tail", "local peak", "negative values"],
        "domains": ["sales", "health", "education", "demographics", "operations", "web analytics"],
        "styles": ["standard", "long labels", "legend at bottom", "data labels"],
    },
    {
        "code": "C03",
        "family": "Line",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C03.01", "Line", 220, False),
            ("C03.02", "Line with Markers", 170, False),
            ("C03.03", "Stacked Line", 110, False),
            ("C03.04", "Stacked Line with Markers", 100, False),
            ("C03.05", "100% Stacked Line", 90, False),
        ],
        "structures": ["time series", "multi-series temporal", "ordered categorical"],
        "tasks": ["read value", "compare series", "find max/min", "detect trend", "detect regime change", "detect seasonality"],
        "patterns": ["growth", "decline", "seasonality", "regime change", "high noise", "sudden spike", "sudden valley"],
        "domains": ["sales", "finance", "energy", "climate", "web analytics", "operations"],
        "styles": ["standard", "with markers", "no grid", "legend on the right"],
    },
    {
        "code": "C04",
        "family": "Pie",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C04.01", "Pie", 80, False),
            ("C04.02", "Pie of Pie", 40, False),
            ("C04.03", "Bar of Pie", 30, False),
            ("C04.04", "Exploded Pie", 30, False),
        ],
        "structures": ["simple part-of-whole", "part-of-whole with long tail"],
        "tasks": ["compare proportions", "identify dominant segment", "estimate share", "detect long tail"],
        "patterns": ["balanced parts", "dominant part", "long tail", "very close segments"],
        "domains": ["sales", "web analytics", "demographics", "education", "operations"],
        "styles": ["standard", "with percentages", "legend on the right", "highlighted segment"],
    },
    {
        "code": "C05",
        "family": "Doughnut",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [("C05.01", "Doughnut", 110, False)],
        "structures": ["simple part-of-whole", "multi-ring part-of-whole"],
        "tasks": ["compare proportions", "identify dominant segment", "compare rings"],
        "patterns": ["balanced parts", "dominant part", "very close segments"],
        "domains": ["sales", "web analytics", "operations", "demographics"],
        "styles": ["standard", "with percentages", "double ring", "highlighted segment"],
    },
    {
        "code": "C06",
        "family": "Area",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C06.01", "Area", 70, False),
            ("C06.02", "Stacked Area", 50, False),
            ("C06.03", "100% Stacked Area", 30, False),
        ],
        "structures": ["time series", "cumulative multi-series temporal"],
        "tasks": ["detect trend", "compare magnitudes", "analyze accumulation", "compare composition over time"],
        "patterns": ["growth", "decline", "seasonality", "regime change", "growing accumulation"],
        "domains": ["energy", "climate", "web analytics", "operations", "sales"],
        "styles": ["standard", "with transparency", "legend on the right", "no grid"],
    },
    {
        "code": "C07",
        "family": "Scatter",
        "block": "Nuclear",
        "modern": False,
        "subtypes": [
            ("C07.01", "Scatter with Only Markers", 120, False),
            ("C07.02", "Scatter with Smooth Lines", 70, False),
            ("C07.03", "Scatter with Smooth Lines and Markers", 90, False),
            ("C07.04", "Scatter with Straight Lines", 60, False),
        ],
        "structures": ["numeric bivariate"],
        "tasks": ["detect correlation", "detect outlier", "detect clusters", "compare points", "detect non-linear relationship"],
        "patterns": ["positive correlation", "negative correlation", "no correlation", "clear outlier", "clusters", "curved relationship"],
        "domains": ["health", "energy", "climate", "demographics", "manufacturing", "education"],
        "styles": ["standard", "with trend line", "large markers", "light grid"],
    },
    {
        "code": "C09",
        "family": "Stock",
        "block": "Specialized",
        "modern": False,
        "subtypes": [
            ("C09.01", "High-Low-Close", 50, False),
            ("C09.02", "Open-High-Low-Close", 50, False),
            ("C09.03", "Volume-High-Low-Close", 40, False),
            ("C09.04", "Volume-Open-High-Low-Close", 50, False),
        ],
        "structures": ["temporal OHLC", "OHLC + volume"],
        "tasks": ["find max/min", "compare open and close", "detect volatility", "detect trend"],
        "patterns": ["high volatility", "low volatility", "upward trend", "downward trend", "single gap"],
        "domains": ["finance", "energy"],
        "styles": ["standard", "visible volume", "no grid", "legend on the right"],
    },
    {
        "code": "C10",
        "family": "Surface",
        "block": "Structured",
        "modern": False,
        "subtypes": [
            ("C10.01", "Contour", 10, False),
            ("C10.02", "Wireframe Contour", 10, False),
        ],
        "structures": ["2D numeric grid"],
        "tasks": ["find maximum peak", "find minimum valley", "detect gradient", "compare regions", "identify ridges"],
        "patterns": ["smooth surface", "rough surface", "central peak", "double peak", "diagonal gradient"],
        "domains": ["climate", "energy", "manufacturing"],
        "styles": ["standard", "visible wireframe", "contour lines", "accent palette"],
    },
    {
        "code": "C11",
        "family": "Radar",
        "block": "Structured",
        "modern": False,
        "subtypes": [("C11.01", "Radar", 30, False), ("C11.02", "Radar with Markers", 30, False), ("C11.03", "Filled Radar", 20, False)],
        "structures": ["multivariable on common axes"],
        "tasks": ["compare profiles", "identify dominant dimension", "compare series", "detect balanced/unbalanced profile"],
        "patterns": ["balanced profile", "spiky profile", "two contrasting series", "one dominant dimension"],
        "domains": ["education", "health", "operations", "manufacturing"],
        "styles": ["standard", "with markers", "filled", "legend at bottom"],
    },
    {
        "code": "C14",
        "family": "Histogram",
        "block": "Specialized",
        "modern": True,
        "subtypes": [("C14.01", "Histogram", 250, False)],
        "structures": ["univariate distribution"],
        "tasks": ["characterize distribution", "detect asymmetry", "detect concentration", "detect outlier", "compare dispersion"],
        "patterns": ["symmetric", "asymmetric", "bimodal", "long tail", "clear outlier"],
        "domains": ["health", "education", "demographics", "energy", "manufacturing"],
        "styles": ["standard", "more bins", "fewer bins", "no grid"],
    },
    {
        "code": "C15",
        "family": "Pareto",
        "block": "Specialized",
        "modern": True,
        "subtypes": [("C15.01", "Pareto", 100, False)],
        "structures": ["ordered categorical + cumulative"],
        "tasks": ["identify dominant categories", "find 80/20 threshold", "order/rank", "interpret cumulative"],
        "patterns": ["clear 80/20", "long tail", "moderate concentration", "strong concentration"],
        "domains": ["operations", "manufacturing", "sales", "web analytics"],
        "styles": ["standard", "highlighted cumulative line", "data labels", "no grid"],
    },
    {
        "code": "C18",
        "family": "Funnel",
        "block": "Specialized",
        "modern": True,
        "subtypes": [("C18.01", "Funnel", 100, False)],
        "structures": ["stage-by-stage process"],
        "tasks": ["identify bottleneck", "calculate largest drop", "compare stages", "interpret conversion"],
        "patterns": ["uniform drop", "clear bottleneck", "late drop", "strong initial drop"],
        "domains": ["web analytics", "sales", "operations", "health"],
        "styles": ["standard", "visible percentages", "labeled stages", "accent palette"],
    },
    {
        "code": "C19",
        "family": "Combo",
        "block": "Nuclear",
        "modern": True,
        "subtypes": [
            ("C19.01", "Clustered Column + Line", 80, False),
            ("C19.02", "Column + Line on Secondary Axis", 60, False),
            ("C19.03", "Custom Combo", 40, False),
        ],
        "structures": ["mixed multi-series", "temporal with secondary axis"],
        "tasks": ["compare series", "detect relationship between volume and rate", "detect trend", "interpret secondary axis"],
        "patterns": ["growth with rate", "volume and percentage", "two different scales", "main series + target"],
        "domains": ["sales", "finance", "web analytics", "energy", "operations"],
        "styles": ["standard", "visible secondary axis", "legend on the right", "selective labels"],
    },
]


def difficulty_sequence(count: int, block: str) -> List[str]:
    if block == "Nuclear":
        ratios = [0.35, 0.40, 0.20, 0.05]
    elif block == "Specialized":
        ratios = [0.20, 0.35, 0.30, 0.15]
    else:
        ratios = [0.15, 0.35, 0.35, 0.15]

    labels = ["low", "medium", "high", "very high"]
    raw = [count * r for r in ratios]
    base = [math.floor(x) for x in raw]

    while sum(base) < count:
        remainders = [raw[i] - base[i] for i in range(4)]
        idx = max(range(4), key=lambda i: (remainders[i], -i))
        base[idx] += 1

    seq: List[str] = []
    pools = {lab: n for lab, n in zip(labels, base)}
    order = ["medium", "low", "high", "medium", "low", "high", "very high"]

    while len(seq) < count:
        for lab in order:
            if pools[lab] > 0 and len(seq) < count:
                seq.append(lab)
                pools[lab] -= 1
    return seq


def question_template(task: str) -> Tuple[str, str]:
    mapping = {
        "read value": ("What value does {element_objetiu} have in the chart?", "numeric_value"),
        "compare categories": ("Which category is larger between {element_A} and {element_B}?", "categorical_comparison"),
        "compare series": ("Which series is higher at point/category {element_objetiu}?", "series_comparison"),
        "find max/min": ("Which element shows the maximum or minimum value?", "extreme_element"),
        "order/rank": ("What are the top elements in descending order?", "ranking"),
        "estimate difference": ("What is the approximate difference between {element_A} and {element_B}?", "difference"),
        "detect trend": ("Is the overall trend increasing, decreasing, or stable?", "trend_pattern"),
        "detect regime change": ("At what point does the series behavior clearly change?", "change_point"),
        "detect seasonality": ("Is there a seasonal or cyclical pattern?", "seasonal_pattern"),
        "compare proportions": ("Which segment represents a larger proportion of the total?", "dominant_segment"),
        "identify dominant segment": ("Which segment dominates the total?", "dominant_segment"),
        "estimate share": ("What approximate share does {element_objetiu} represent?", "percentage"),
        "detect long tail": ("Is there a long tail of small segments?", "composition_pattern"),
        "compare magnitudes": ("Which series or interval has the highest magnitude?", "magnitude_comparison"),
        "analyze accumulation": ("How does the cumulative total evolve across the series?", "accumulation"),
        "compare composition over time": ("How does the relative composition change over time?", "temporal_composition"),
        "detect correlation": ("Is there a positive, negative, or no correlation?", "correlation"),
        "detect outlier": ("Is there any clearly anomalous point or value?", "outlier"),
        "detect outliers": ("Is there any visible atypical value or outlier?", "outlier"),
        "detect clusters": ("Are there distinct groupings or clusters?", "clusters"),
        "compare points": ("Which point or group of points stands out the most?", "point_comparison"),
        "detect non-linear relationship": ("Does the relationship look linear or curved?", "relationship_shape"),
        "compare bubble sizes": ("Which bubble represents the largest size?", "relative_size"),
        "compare open and close": ("In which period is the opening-closing difference the largest?", "open_close_difference"),
        "detect volatility": ("In which segment is volatility highest?", "volatility"),
        "find maximum peak": ("Where is the maximum peak of the surface?", "extreme_region"),
        "find minimum valley": ("Where is the minimum valley of the surface?", "extreme_region"),
        "detect gradient": ("Which direction shows the main gradient?", "gradient"),
        "compare regions": ("Which region or area has the highest value?", "extreme_region"),
        "identify ridges": ("Are ridges or distinct elevated zones identifiable?", "surface_morphology"),
        "compare profiles": ("Which profile is more balanced or more extreme?", "profile_comparison"),
        "identify dominant dimension": ("Which dimension or axis dominates the profile?", "dominant_dimension"),
        "detect balanced/unbalanced profile": ("Is the profile balanced or dominated by a few axes?", "overall_profile"),
        "identify dominant branch": ("Which hierarchical branch accumulates the most weight?", "dominant_branch"),
        "compare parts of the whole": ("Which subgroup represents a larger part of the total?", "parts_comparison"),
        "detect hierarchy": ("What is the chart's main hierarchical structure?", "hierarchy"),
        "compare subgroups": ("Which subgroup is larger within the main branch?", "subgroup_comparison"),
        "compare hierarchical rings": ("Which hierarchical level concentrates the most value?", "dominant_level"),
        "compare rings": ("Which ring level concentrates the most value?", "dominant_level"),
        "characterize distribution": ("Is the distribution concentrated, dispersed, symmetric, or asymmetric?", "distribution_description"),
        "detect asymmetry": ("Is the distribution symmetric or does it have a tail on one side?", "asymmetry"),
        "detect concentration": ("Are most values concentrated within a narrow range?", "concentration"),
        "compare dispersion": ("Which group or distribution shows greater dispersion?", "dispersion"),
        "identify dominant categories": ("Which categories concentrate most of the impact?", "dominant_categories"),
        "find 80/20 threshold": ("How many categories are needed to reach approximately 80% cumulative?", "threshold_80_20"),
        "interpret cumulative": ("How does the cumulative line or total evolve?", "accumulation"),
        "compare medians": ("Which group has the highest median?", "median"),
        "compare asymmetry": ("Which group shows more asymmetry?", "asymmetry"),
        "identify main contribution": ("Which step contributes the most to the final result?", "dominant_step"),
        "find most positive/negative step": ("Which step has the largest positive or negative effect?", "extreme_step"),
        "compare initial and final balance": ("How does the total change from start to end?", "total_change"),
        "identify bottleneck": ("At which stage does the largest loss occur?", "bottleneck"),
        "calculate largest drop": ("Between which stages is the largest drop?", "maximum_drop"),
        "compare stages": ("Which stage retains the highest relative volume?", "stage_comparison"),
        "interpret conversion": ("What is the approximate conversion to the final stage?", "conversion"),
        "detect relationship between volume and rate": ("Do the line and columns show aligned or divergent behavior?", "series_relationship"),
        "interpret secondary axis": ("Which series is interpreted on the secondary axis and what does it imply?", "secondary_axis"),
        "detect hotspot": ("Is there any region acting as a clear hotspot?", "hotspot"),
        "detect spatial gradient": ("Is there a recognizable spatial gradient?", "spatial_gradient"),
        "identify extreme region": ("Which region shows the primary extreme?", "extreme_region"),
    }
    return mapping.get(task, ("What is the main takeaway from this chart?", "qualitative_answer"))


def style_sequence(count: int, styles: List[str], family: str | None = None) -> List[str]:
    base = list(styles)
    return [base[i % len(base)] for i in range(count)]


def family_summary(families: List[Dict]) -> List[Dict]:
    rows = []
    for fam in families:
        q5000 = sum(st[2] for st in fam["subtypes"])
        rows.append(
            {
                "family_code": fam["code"],
                "Excel_family": fam["family"],
                "coverage_block": fam["block"],
                "n_subtype": len(fam["subtypes"]),
                "is_modern": "Yes" if fam["modern"] else "No",
                "quote_5000": q5000,
                "quote_500": q5000 // 10,
                "weight_5000": q5000 / 5000,
                "weight_500": (q5000 // 10) / 500,
            }
        )
    return rows


def generate_cases(families: List[Dict], seed: int = RANDOM_SEED, target_full: int = 5000 , target_sample: int = 500) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    rng = random.Random(seed)
    full_cases: List[Dict] = []
    coverage_rows: List[Dict] = []
    case_num = 1
    
    # Calculate base total from family quotas
    base_total = sum(q for fam in families for _, _, q, _ in fam["subtypes"])
    scale_factor = target_full / base_total if base_total > 0 else 1.0

    # First pass: calculate scaled quotas
    quota_map = {}
    for fam in families:
        for subtype_code, subtype, quota_base, is_3d in fam["subtypes"]:
            quota_full = max(1, round(quota_base * scale_factor))
            quota_sample = max(1, round(quota_full * target_sample / target_full))
            quota_map[(fam["code"], subtype_code)] = {
                "quota_full": quota_full,
                "quota_sample": quota_sample,
                "quota_base": quota_base
            }
    
    # Adjust to hit exact target
    current_total = sum(q["quota_full"] for q in quota_map.values())
    diff = target_full - current_total
    
    if diff != 0:
        # Sort by base quota (adjust largest items first)
        sorted_keys = sorted(quota_map.keys(), key=lambda k: quota_map[k]["quota_base"], reverse=(diff > 0))
        for i, key in enumerate(sorted_keys):
            if diff == 0:
                break
            adjustment = 1 if diff > 0 else -1
            if quota_map[key]["quota_full"] + adjustment >= 1:
                quota_map[key]["quota_full"] += adjustment
                # Proportionally adjust sample quota
                new_sample = max(1, round(quota_map[key]["quota_full"] * target_sample / target_full))
                quota_map[key]["quota_sample"] = new_sample
                diff -= adjustment

    # Second pass: ensure sample total is correct
    current_sample_total = sum(q["quota_sample"] for q in quota_map.values())
    sample_diff = target_sample - current_sample_total
    
    if sample_diff != 0:
        sorted_keys = sorted(quota_map.keys(), key=lambda k: quota_map[k]["quota_base"], reverse=(sample_diff > 0))
        for key in sorted_keys:
            if sample_diff == 0:
                break
            adjustment = 1 if sample_diff > 0 else -1
            if quota_map[key]["quota_sample"] + adjustment >= 1:
                quota_map[key]["quota_sample"] += adjustment
                sample_diff -= adjustment

    for fam in families:
        for subtype_code, subtype, quota_base, is_3d in fam["subtypes"]:
            quotas = quota_map[(fam["code"], subtype_code)]
            quota_full = quotas["quota_full"]
            quota_sample = quotas["quota_sample"]
            difficulties = difficulty_sequence(quota_full, fam["block"])
            styles = style_sequence(quota_full, fam["styles"], fam["family"])

            structures = [fam["structures"][(idx * 2 + 1) % len(fam["structures"])] for idx in range(quota_full)]
            tasks = [fam["tasks"][(idx * 3 + 2) % len(fam["tasks"])] for idx in range(quota_full)]
            patterns = [fam["patterns"][(idx * 5 + 1) % len(fam["patterns"])] for idx in range(quota_full)]
            domains = [fam["domains"][(idx * 7 + 3) % len(fam["domains"])] for idx in range(quota_full)]

            if subtype in {"Pie of Pie", "Bar of Pie"}:
                structures = ["parts of total with long tail"] * quota_full
                patterns = ["long tail"] * quota_full
            if fam["family"] == "Surface":
                structures = ["numeric grid 2D"] * quota_full
            if fam["family"] == "Histogram":
                structures = ["univariate distribution"] * quota_full

            selected_orders = set(sorted(rng.sample(range(1, quota_full + 1), quota_sample)))

            coverage_rows.append(
                {
                    "family_code": fam["code"],
                    "Excel_family": fam["family"],
                    "code_subtype": subtype_code,
                    "subtype_excel": subtype,
                    "coverage_block": fam["block"],
                    "is_modern": "Yes" if fam["modern"] else "No",
                    "is_3d": "Yes" if is_3d else "No",
                    "main_structure": ", ".join(fam["structures"]),
                    "compatible_tasks": ", ".join(fam["tasks"]),
                    "compatible_patterns": ", ".join(fam["patterns"]),
                    "quote_5000": quota_full,
                    "quote_500": quota_sample,
                    "weight_5000": quota_full / target_full,
                    "weight_500": quota_sample / target_sample,
                }
            )

            for idx in range(quota_full):
                task = tasks[idx]
                prompt, answer_type = question_template(task)
                order = idx + 1
                full_cases.append(
                    {
                        "case_id": f"CASE_{case_num:05d}",
                        "family_code": fam["code"],
                        "Excel_family": fam["family"],
                        "code_subtype": subtype_code,
                        "subtype_excel": subtype,
                        "coverage_block": fam["block"],
                        "data_structure": structures[idx],
                        "statistic_pattern": patterns[idx],
                        "analytic_task": task,
                        "question_template": prompt,
                        "answer_type_oracle": answer_type,
                        "difficulty": difficulties[idx],
                        "semantic_domain": domains[idx],
                        "style_varian": styles[idx],
                        "is_modern": "Yes" if fam["modern"] else "No",
                        "quote_subtype_5000": quota_full,
                        "quote_subtype_500": quota_sample,
                        "order_within_subtype": order,
                        "selected_in_sample_500": "Yes" if order in selected_orders else "No",
                    }
                )
                case_num += 1

    sample_cases = [row for row in full_cases if row["selected_in_sample_500"] == "Yes"]
    return full_cases, sample_cases, coverage_rows


# ---------- Excel helpers ----------
fill_title = PatternFill("solid", fgColor="1F4E78")
fill_subheader = PatternFill("solid", fgColor="EAF3F8")
fill_nuclear = PatternFill("solid", fgColor="DDEBF7")
fill_special = PatternFill("solid", fgColor="FFF2CC")
fill_struct = PatternFill("solid", fgColor="E4DFEC")
fill_modern = PatternFill("solid", fgColor="E2F0D9")
fill_3d = PatternFill("solid", fgColor="FCE4D6")
white_font = Font(color="FFFFFF", bold=True, size=12)
title_font = Font(color="FFFFFF", bold=True, size=14)
header_font = Font(bold=True, color="1F1F1F")
normal_font = Font(color="000000", size=10)
blue_font = Font(color="1F4E78", bold=True)
border_light = Border(bottom=Side(style="thin", color="D9D9D9"))


def add_sheet(wb: Workbook, name: str, headers: List[str], rows: List[Dict], widths: Dict[int, int] | None = None, style_rows: bool = True):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill_title
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_light

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.sheet_view.showGridLines = False

    if style_rows:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = normal_font
                cell.border = border_light

    if widths:
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    return ws


def build_workbook(full_cases: List[Dict], sample_cases: List[Dict], coverage_rows: List[Dict], output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Aggregate family summary from actual scaled coverage rows
    fam_seen: Dict[str, Dict] = {}
    for row in coverage_rows:
        fc = row["family_code"]
        if fc not in fam_seen:
            fam_seen[fc] = {
                "family_code": fc,
                "Excel_family": row["Excel_family"],
                "coverage_block": row["coverage_block"],
                "is_modern": row["is_modern"],
                "n_subtype": 0,
                "quote_5000": 0,
                "quote_500": 0,
            }
        fam_seen[fc]["n_subtype"] += 1
        fam_seen[fc]["quote_5000"] += row["quote_5000"]
        fam_seen[fc]["quote_500"] += row["quote_500"]
    fam_summary = list(fam_seen.values())

    # Overview sheet
    ws = wb.create_sheet("Overview")
    ws["A1"] = "Coverage Matrix for Excel Charts"
    ws["A1"].fill = fill_title
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")

    lines = [
        ("Objective", "Define comprehensive coverage of Excel chart types with a complete matrix of 5,000 cases and an initial proportional sample of 500."),
        ("Sampling criteria", "The sample of 500 is exactly 10% of each subtype from the 5,000 matrix."),
        ("Selection method", "Stratified and deterministic random selection within each subtype (fixed seed)."),
        ("Coverage level", "Coverage is organized by family, subtype, data structure, pattern, task, difficulty, and semantic domain."),
        ("Subtypes", "60 operational subtypes have been modeled."),
    ]
    row = 3
    for label, text in lines:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = blue_font
        ws[f"A{row}"].fill = fill_subheader
        ws[f"B{row}"] = text
        row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    ws.sheet_view.showGridLines = False

    # Family summary
    fam_headers = [
        "family_code", "Excel_family", "coverage_block", "n_subtype", "is_modern",
        "quote_goal_5000", "real_5000", "quote_goal_500", "real_500",
        "weight_5000", "weight_500",
    ]
    wsf = wb.create_sheet("Family_Summary")
    wsf.append(fam_headers)
    for c in wsf[1]:
        c.fill = fill_title
        c.font = white_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border_light

    for i, fam in enumerate(fam_summary, start=2):
        wsf[f"A{i}"] = fam["family_code"]
        wsf[f"B{i}"] = fam["Excel_family"]
        wsf[f"C{i}"] = fam["coverage_block"]
        wsf[f"D{i}"] = fam["n_subtype"]
        wsf[f"E{i}"] = fam["is_modern"]
        wsf[f"F{i}"] = fam["quote_5000"]
        wsf[f"G{i}"] = f'=COUNTIF(Matrix_5000!B:B,A{i})'
        wsf[f"H{i}"] = fam["quote_500"]
        wsf[f"I{i}"] = f'=COUNTIF(Matrix_500!B:B,A{i})'
        wsf[f"J{i}"] = f'=G{i}/5000'
        wsf[f"K{i}"] = f'=I{i}/500'

    wsf.freeze_panes = "A2"
    wsf.auto_filter.ref = f"A1:K{wsf.max_row}"
    wsf.sheet_view.showGridLines = False

    # Subtype summary
    sub_headers = [
        "family_code", "Excel_family", "code_subtype", "subtype_excel", "coverage_block", "is_modern",
        "is_3d", "main_structure", "compatible_tasks", "compatible_patterns",
        "quote_goal_5000", "actual_5000", "quote_goal_500", "actual_500", "weight_5000", "weight_500",
    ]
    wss = wb.create_sheet("Subtype_Summary")
    wss.append(sub_headers)
    for c in wss[1]:
        c.fill = fill_title
        c.font = white_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border_light

    for i, rowd in enumerate(coverage_rows, start=2):
        wss[f"A{i}"] = rowd["family_code"]
        wss[f"B{i}"] = rowd["Excel_family"]
        wss[f"C{i}"] = rowd["code_subtype"]
        wss[f"D{i}"] = rowd["subtype_excel"]
        wss[f"E{i}"] = rowd["coverage_block"]
        wss[f"F{i}"] = rowd["is_modern"]
        wss[f"G{i}"] = rowd["is_3d"]
        wss[f"H{i}"] = rowd["main_structure"]
        wss[f"I{i}"] = rowd["compatible_tasks"]
        wss[f"J{i}"] = rowd["compatible_patterns"]
        wss[f"K{i}"] = rowd["quote_5000"]
        wss[f"L{i}"] = f'=COUNTIF(Matrix_5000!D:D,C{i})'
        wss[f"M{i}"] = rowd["quote_500"]
        wss[f"N{i}"] = f'=COUNTIF(Matrix_500!D:D,C{i})'
        wss[f"O{i}"] = f'=L{i}/5000'
        wss[f"P{i}"] = f'=N{i}/500'

    wss.freeze_panes = "A2"
    wss.auto_filter.ref = f"A1:P{wss.max_row}"
    wss.sheet_view.showGridLines = False

    # Matrices
    matrix_headers = list(full_cases[0].keys())
    add_sheet(
        wb,
        "Matrix_5000",
        matrix_headers,
        full_cases,
        widths={1: 12, 2: 10, 3: 14, 4: 12, 5: 28, 6: 16, 7: 18, 8: 22, 9: 28, 10: 48, 11: 20, 12: 12, 13: 16, 14: 18, 15: 10, 16: 8, 17: 16, 18: 16, 19: 16, 20: 14},
        style_rows=False,
    )
    add_sheet(
        wb,
        "Matrix_500",
        matrix_headers,
        sample_cases,
        widths={1: 12, 2: 10, 3: 14, 4: 12, 5: 28, 6: 16, 7: 18, 8: 22, 9: 28, 10: 48, 11: 20, 12: 12, 13: 16, 14: 18, 15: 10, 16: 8, 17: 16, 18: 16, 19: 16, 20: 14},
        style_rows=False,
    )

    # Colors per block and flags
    # Colors per block and flags (summary sheets only — matrix sheets are too large for per-cell styling)
    for sheet_name, block_col, modern_col, d3_col, cols in [
        ("Family_Summary",   3, 5, None, [1, 2, 3]),
        ("Subtype_Summary",  5, 6, 7,    [1, 2, 3, 4]),
    ]:
        wsx = wb[sheet_name]
        for r in range(2, wsx.max_row + 1):
            block = wsx.cell(r, block_col).value
            fill = fill_nuclear if block == "Nuclear" else fill_special if block == "Specialized" else fill_struct
            for c in cols:
                wsx.cell(r, c).fill = fill
            if wsx.cell(r, modern_col).value == "Yes":
                wsx.cell(r, modern_col).fill = fill_modern
            if d3_col and wsx.cell(r, d3_col).value == "Yes":
                wsx.cell(r, d3_col).fill = fill_3d

    wb.save(output_path)

### ---------- Main execution ----------

###
def main(target_full_cases: int = 5000, target_sample_cases: int = 500) -> None:
    full_cases, sample_cases, coverage_rows = generate_cases(FAMILIES, seed=RANDOM_SEED, target_full=target_full_cases, target_sample=target_sample_cases)

    assert len(full_cases) == target_full_cases, f"Expected {target_full_cases} cases but got {len(full_cases)}"
    assert len(sample_cases) == target_sample_cases, f"Expected {target_sample_cases} cases but got {len(sample_cases)}"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_xlsx = OUTPUT_DIR / f"coverage_matrix_{target_full_cases}_{target_sample_cases}.xlsx"
    output_csv_full = OUTPUT_DIR / f"matrix_{target_full_cases}.csv"
    output_csv_sample = OUTPUT_DIR / f"matrix_{target_sample_cases}.csv"

    build_workbook(full_cases, sample_cases, coverage_rows, output_xlsx)
    pd.DataFrame(full_cases).to_csv(output_csv_full, index=False, encoding="utf-8-sig")
    pd.DataFrame(sample_cases).to_csv(output_csv_sample, index=False, encoding="utf-8-sig")

    print(f"Excel file created: {output_xlsx.resolve()}")
    print(f"CSV {target_full_cases} created:   {output_csv_full.resolve()}")
    print(f"CSV {target_sample_cases} created:    {output_csv_sample.resolve()}")


if __name__ == "__main__":
    main()


