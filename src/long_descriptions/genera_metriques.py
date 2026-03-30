from __future__ import annotations

import csv
import json
import math
import re
from pathlib import Path
from functools import lru_cache
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook
import pronouncing
import pyphen


BASE_DIR = Path.cwd()
LONG_DESCRIPTIONS_DIR = BASE_DIR / "long-descriptions"
OUTPUT_DIR = BASE_DIR / "sortida_instancies_completa"
METRICS_DIR = LONG_DESCRIPTIONS_DIR / "metriques"

# Configuració de proveedors
PROVIDERS = {
	"claude": LONG_DESCRIPTIONS_DIR / "claude" / "claude_500_casos.xlsx",
	"gemini": LONG_DESCRIPTIONS_DIR / "gemini" / "gemini_500_casos.xlsx",
	"chatgpt": LONG_DESCRIPTIONS_DIR / "chatgpt" / "chatgpt_500_casos.xlsx",
}

OUTPUT_CSV_UNIFIED = METRICS_DIR / "metriques_unificat.csv"
OUTPUT_JSON_UNIFIED = METRICS_DIR / "metriques_unificat.json"
OUTPUT_CSV_AGGREGATED = METRICS_DIR / "metriques_agregat.csv"
OUTPUT_JSON_AGGREGATED = METRICS_DIR / "metriques_agregat.json"
CANONICAL_JSON = OUTPUT_DIR / "instancies_canoniques.json"

# Para backwards compatibility
OUTPUT_CSV = METRICS_DIR / "metriques_claude_paraules_numeros.csv"
OUTPUT_JSON = METRICS_DIR / "metriques_claude_paraules_numeros.json"

WORD_PATTERN = re.compile(r"[A-Za-z]+(?:[-'][A-Za-z]+)*")
NUMERIC_PATTERN = re.compile(r"(?<![A-Za-z])[+-]?(?:\d+[\.,]?\d*|\d*[\.,]\d+)(?![A-Za-z])")
CASE_ID_PATTERN = re.compile(r"CASE_\d+")
CASE_ID_IN_TEXT_PATTERN = re.compile(r"\bcase_?\d+\b|\(0*\d+\)", re.IGNORECASE)
# Split sentences on '.', '!', '?' but keep decimal numbers intact (e.g., 38.6)
SENTENCE_SPLIT_PATTERN = re.compile(r"(?:\.(?!\d)|[!?])+")
HYPHENATOR = pyphen.Pyphen(lang="en_US")

CARDINAL_UNITS = {
	"zero",
	"one",
	"two",
	"three",
	"four",
	"five",
	"six",
	"seven",
	"eight",
	"nine",
	"ten",
	"eleven",
	"twelve",
	"thirteen",
	"fourteen",
	"fifteen",
	"sixteen",
	"seventeen",
	"eighteen",
	"nineteen",
}
CARDINAL_TENS = {
	"twenty",
	"thirty",
	"forty",
	"fifty",
	"sixty",
	"seventy",
	"eighty",
	"ninety",
}
CARDINAL_SCALES = {"hundred", "thousand", "million", "billion", "trillion"}
ORDINAL_WORDS = {
	"first",
	"second",
	"third",
	"fourth",
	"fifth",
	"sixth",
	"seventh",
	"eighth",
	"ninth",
	"tenth",
	"eleventh",
	"twelfth",
	"thirteenth",
	"fourteenth",
	"fifteenth",
	"sixteenth",
	"seventeenth",
	"eighteenth",
	"nineteenth",
	"twentieth",
	"thirtieth",
	"fortieth",
	"fiftieth",
	"sixtieth",
	"seventieth",
	"eightieth",
	"ninetieth",
	"hundredth",
	"thousandth",
	"millionth",
	"billionth",
	"trillionth",
}
CARDINAL_NUMBER_WORDS = CARDINAL_UNITS | CARDINAL_TENS | CARDINAL_SCALES

QCA_TOLERANCE_REL = 0.01
QCA_TOLERANCE_ABS = 0.1
QCA_APPROX_TOLERANCE_REL = 0.025
QCA_APPROX_TOLERANCE_ABS = 0.2

QCA_APPROX_TERMS = {
	"about",
	"approximately",
	"approx",
	"around",
	"roughly",
	"nearly",
	"close to",
}

QCA_TREND_PATTERNS = {
	"increase_then_decrease": ["increase", "decrease"],
	"decrease_then_increase": ["decrease", "increase"],
	"increasing": ["increase", "rising", "upward", "growth"],
	"decreasing": ["decrease", "decline", "fall", "downward"],
	"stable": ["stable", "flat", "constant", "unchanged"],
}


def extract_case_id(chart_label: str) -> str:
	match = CASE_ID_PATTERN.search(chart_label or "")
	if match:
		return match.group(0)
	return chart_label.strip()


def normalize_spaces(text: str) -> str:
	return re.sub(r"\s+", " ", text).strip()


def tokenize_words(text: str) -> List[str]:
	return WORD_PATTERN.findall(text)


def count_numeric_tokens(text: str) -> int:
	return len(NUMERIC_PATTERN.findall(text))


def extract_numeric_values(text: str) -> List[float]:
	values: List[float] = []
	for token in NUMERIC_PATTERN.findall(text):
		normalized = token.replace(",", ".")
		try:
			values.append(float(normalized))
		except ValueError:
			continue
	return values


def normalize_category_numeric_token(token: str) -> str:
	token = token.strip().replace(",", ".")
	try:
		value = float(token)
	except ValueError:
		return token
	if value.is_integer():
		return str(int(value))
	return str(value)


def is_x_axis_category_number(value: float, canonical_case: Dict[str, Any] | None) -> bool:
	if not canonical_case:
		return False
	categories = canonical_case.get("categories") or set()
	if not categories:
		return False
	if value.is_integer():
		as_int = str(int(value))
		if as_int in categories:
			return True
	return normalize_category_numeric_token(str(value)) in categories


def filter_value_numbers(numbers: List[float], canonical_case: Dict[str, Any] | None) -> List[float]:
	"""Filter out numeric mentions that correspond to x-axis numeric categories (e.g., years)."""
	if not canonical_case:
		return numbers
	return [number for number in numbers if not is_x_axis_category_number(number, canonical_case)]


def filter_non_data_numbers(sentence: str, numbers: List[float]) -> List[float]:
	"""Filter numbers that are likely metadata/context rather than chart data values."""
	if not numbers:
		return numbers
	lower = sentence.lower()

	# Case IDs in titles, e.g. (00024)
	if CASE_ID_IN_TEXT_PATTERN.search(lower):
		return []

	# Aggregated/statistical references usually are not point values in canonical series.
	if re.search(r"\b(average|mean|median)\b", lower):
		return []

	# Percentages are typically derived values, not raw y-series points.
	if "%" in sentence or re.search(r"\b(percent|percentage)\b", lower):
		return []

	# Metadata counters (series/categories/points/total/count) are not y-values.
	if re.search(r"\b(series|categories|category|points?|groups?)\b", lower):
		return []
	if re.search(r"\b(total|count|number of)\b", lower) and re.search(r"\b(cases|series|categories|points?)\b", lower):
		return []

	return numbers


def is_temporal_range_only_claim(sentence: str, raw_numbers: List[float], filtered_numbers: List[float]) -> bool:
	"""True when a structural range claim only references years (timeline, not data values)."""
	if filtered_numbers:
		return False
	if not raw_numbers:
		return False
	lower = sentence.lower()
	if not re.search(r"\b(from|between|range|ranges|ranging|spans)\b", lower):
		return False
	return all(number.is_integer() and 1900 <= int(number) <= 2100 for number in raw_numbers)


def count_sentences(text: str) -> int:
	parts = [part.strip() for part in SENTENCE_SPLIT_PATTERN.split(text) if part.strip()]
	return len(parts)


@lru_cache(maxsize=50000)
def estimate_syllables_in_word(word: str) -> int:
	clean_word = re.sub(r"[^a-z]", "", word.lower())
	if not clean_word:
		return 0

	phones = pronouncing.phones_for_word(clean_word)
	if phones:
		return max(1, pronouncing.syllable_count(phones[0]))

	hyphenated = HYPHENATOR.inserted(clean_word)
	if hyphenated:
		parts = [part for part in hyphenated.split("-") if part]
		if parts:
			return max(1, len(parts))

	return 1


@lru_cache(maxsize=50000)
def is_word_in_dictionary(word: str) -> bool:
	clean_word = re.sub(r"[^a-z]", "", word.lower())
	if not clean_word:
		return False
	return bool(pronouncing.phones_for_word(clean_word))


def count_dictionary_confidence(text: str) -> float | None:
	words = tokenize_words(text)
	if not words:
		return None
	in_dict = sum(1 for word in words if is_word_in_dictionary(word))
	return round((in_dict / len(words)) * 100, 6)


def count_total_syllables(text: str) -> int:
	total = 0
	for token in tokenize_words(text):
		parts = [part for part in re.split(r"[-']", token) if part]
		for part in parts:
			total += estimate_syllables_in_word(part)
	return total


def count_characters(text: str) -> int:
	return len(re.sub(r"\s", "", text))


def gunning_fog_index(word_count: int, sentence_count: int, polysyllabic_count: int) -> float | None:
	if word_count == 0 or sentence_count == 0:
		return None
	if word_count < 100:
		return None
	words_per_sentence = word_count / sentence_count
	polysyllabic_ratio = (polysyllabic_count / word_count) * 100
	return round(0.4 * (words_per_sentence + 100 * (polysyllabic_ratio / 100)), 6)


def flesch_kincaid_grade(word_count: int, sentence_count: int, syllable_count: int) -> float | None:
	if word_count == 0 or sentence_count == 0:
		return None
	syllables_per_word = syllable_count / word_count
	words_per_sentence = word_count / sentence_count
	return round(0.39 * words_per_sentence + 11.8 * syllables_per_word - 15.59, 6)


def coleman_liau_index(char_count: int, sentence_count: int, word_count: int) -> float | None:
	if word_count == 0:
		return None
	# L = characters per 100 words, S = sentences per 100 words
	L = (char_count / word_count) * 100
	S = (sentence_count / word_count) * 100
	return round(0.0588 * L - 0.296 * S - 15.8, 6)


def count_polysyllabic_words(text: str) -> int:
	count = 0
	for token in tokenize_words(text):
		parts = [part for part in re.split(r"[-']", token) if part]
		syllables = sum(estimate_syllables_in_word(part) for part in parts)
		if syllables >= 3:
			count += 1
	return count


def smog_index(polysyllabic_word_count: int, sentence_count: int) -> float | None:
	if polysyllabic_word_count <= 0:
		return 0.0
	if sentence_count < 0:
		return None
	return round(1.0430 * math.sqrt(polysyllabic_word_count * (30 / (sentence_count + 3.1291))), 6)


def has_chart_type(text: str) -> bool:
	"""Detect if description mentions chart type."""
	chart_types = r"\b(line|bar|pie|histogram|scatter|area|column|clustered|stacked|bubble|heatmap|waterfall|funnel|treemap|sunburst|radial|gauge|combo|mixed)\s+(chart|plot|graph|diagram)\b"
	return bool(re.search(chart_types, text.lower()))


def has_title(text: str) -> bool:
	"""Detect if description mentions chart title."""
	title_patterns = r"\b(title|titled|named|labeled as|called|shows|represents)\b"
	return bool(re.search(title_patterns, text.lower()))


def has_axis_labels(text: str) -> bool:
	"""Detect if description mentions axis labels."""
	axis_patterns = r"\b(axis|axes|x-axis|y-axis|horizontal|vertical|label|labelled|labeled|represents)\b"
	return bool(re.search(axis_patterns, text.lower()))


def has_categories(text: str) -> bool:
	"""Detect if description mentions categories or groups."""
	category_patterns = r"\b(categor|group|section|region|period|country|product|service|type|kind|shows|lists)\b"
	return bool(re.search(category_patterns, text.lower()))


def has_values(text: str) -> bool:
	"""Detect if description mentions numeric values."""
	return count_numeric_tokens(text) > 0


def has_scale_info(text: str) -> bool:
	"""Detect if description mentions scale or range information."""
	scale_patterns = r"\b(range|scale|spans|ranging|varies|between|from.*to|minimum|maximum|extent|values|thousands|million|percent)\b"
	return bool(re.search(scale_patterns, text.lower()))


def qca_is_close(text_value: float, true_value: float, approximate: bool) -> bool:
	rel_tol = QCA_APPROX_TOLERANCE_REL if approximate else QCA_TOLERANCE_REL
	abs_tol = QCA_APPROX_TOLERANCE_ABS if approximate else QCA_TOLERANCE_ABS
	threshold = max(rel_tol * abs(true_value), abs_tol)
	return abs(text_value - true_value) <= threshold


def qca_sequence_trend(values: List[float]) -> str:
	if len(values) < 3:
		return "mixed"

	diffs = [values[i + 1] - values[i] for i in range(len(values) - 1)]
	positive = sum(1 for diff in diffs if diff > 0)
	negative = sum(1 for diff in diffs if diff < 0)
	zero = len(diffs) - positive - negative

	if zero / len(diffs) >= 0.9:
		return "stable"
	if positive == len(diffs):
		return "increasing"
	if negative == len(diffs):
		return "decreasing"

	peak_index = max(range(len(values)), key=lambda idx: values[idx])
	valley_index = min(range(len(values)), key=lambda idx: values[idx])
	if 0 < peak_index < len(values) - 1:
		left = values[: peak_index + 1]
		right = values[peak_index:]
		if all(left[i + 1] >= left[i] for i in range(len(left) - 1)) and all(
			right[i + 1] <= right[i] for i in range(len(right) - 1)
		):
			return "increase_then_decrease"
	if 0 < valley_index < len(values) - 1:
		left = values[: valley_index + 1]
		right = values[valley_index:]
		if all(left[i + 1] <= left[i] for i in range(len(left) - 1)) and all(
			right[i + 1] >= right[i] for i in range(len(right) - 1)
		):
			return "decrease_then_increase"

	if positive / len(diffs) >= 0.9:
		return "increasing"
	if negative / len(diffs) >= 0.9:
		return "decreasing"
	return "mixed"


def load_canonical_cases() -> Dict[str, Dict[str, Any]]:
	if not CANONICAL_JSON.exists():
		raise FileNotFoundError(f"No s'ha trobat el fitxer canònic: {CANONICAL_JSON}")

	with CANONICAL_JSON.open("r", encoding="utf-8") as handle:
		raw_items = json.load(handle)

	parsed: Dict[str, Dict[str, Any]] = {}
	for item in raw_items:
		case_id = str(item.get("id") or "").strip()
		if not case_id:
			continue

		values_raw = (((item.get("data") or {}).get("source") or {}).get("values") or [])
		numeric_values: List[float] = []
		categories: List[str] = []
		series: List[str] = []

		for point in values_raw:
			value = point.get("valor")
			if value is None:
				value = point.get("y")
			if value is None:
				value = point.get("value")
			if isinstance(value, (int, float)):
				numeric_values.append(float(value))

			category = point.get("categoria")
			if category is None:
				category = point.get("x")
			if category is None:
				category = point.get("data")
			if category is not None:
				categories.append(str(category).strip().lower())

			serie = point.get("serie")
			if serie is None:
				serie = point.get("grup")
			if serie is not None:
				series.append(str(serie).strip().lower())

		axes = item.get("axes") or {}
		y_axis = axes.get("y") or {}
		y_min = y_axis.get("min")
		y_max = y_axis.get("max")
		if not isinstance(y_min, (int, float)):
			y_min = min(numeric_values) if numeric_values else None
		if not isinstance(y_max, (int, float)):
			y_max = max(numeric_values) if numeric_values else None

		parsed[case_id] = {
			"values": numeric_values,
			"value_set": set(round(value, 6) for value in numeric_values),
			"categories": set(cat for cat in categories if cat),
			"series": set(ser for ser in series if ser),
			"min": float(y_min) if isinstance(y_min, (int, float)) else None,
			"max": float(y_max) if isinstance(y_max, (int, float)) else None,
			"trend": qca_sequence_trend(numeric_values),
		}

	return parsed


def evaluate_qca(description_text: str, canonical_case: Dict[str, Any] | None) -> Dict[str, Any]:
	if not canonical_case:
		return {
			"qca_score": None,
			"qca_status": "missing_canonical_case",
			"qca_claims_total": 0,
			"qca_claims_verifiable": 0,
			"qca_claims_correct": 0,
			"qca_claims_incorrect": 0,
			"qca_claims_unverifiable": 0,
			"qca_error_value": 0,
			"qca_error_maxmin": 0,
			"qca_error_comparison": 0,
			"qca_error_range": 0,
			"qca_error_trend": 0,
			"qca_error_mapping": 0,
		}

	values = canonical_case.get("values") or []
	if not values:
		return {
			"qca_score": None,
			"qca_status": "missing_canonical_values",
			"qca_claims_total": 0,
			"qca_claims_verifiable": 0,
			"qca_claims_correct": 0,
			"qca_claims_incorrect": 0,
			"qca_claims_unverifiable": 0,
			"qca_error_value": 0,
			"qca_error_maxmin": 0,
			"qca_error_comparison": 0,
			"qca_error_range": 0,
			"qca_error_trend": 0,
			"qca_error_mapping": 0,
		}

	claims_total = 0
	claims_verifiable = 0
	claims_correct = 0
	claims_incorrect = 0
	claims_unverifiable = 0

	error_value = 0
	error_maxmin = 0
	error_comparison = 0
	error_range = 0
	error_trend = 0
	error_mapping = 0

	canonical_min = canonical_case.get("min")
	canonical_max = canonical_case.get("max")
	canonical_trend = canonical_case.get("trend")

	for sentence in [segment.strip() for segment in SENTENCE_SPLIT_PATTERN.split(description_text) if segment.strip()]:
		lower = sentence.lower()
		raw_numbers = extract_numeric_values(sentence)
		numbers = filter_non_data_numbers(sentence, filter_value_numbers(raw_numbers, canonical_case))
		approximate = any(term in lower for term in QCA_APPROX_TERMS)
		skip_structural = is_temporal_range_only_claim(sentence, raw_numbers, numbers)

		if re.search(r"\b(highest|maximum|max|lowest|minimum|min)\b", lower) and not skip_structural:
			claims_total += 1
			if not numbers:
				claims_unverifiable += 1
				error_mapping += 1
			else:
				claims_verifiable += 1
				ref = canonical_max if re.search(r"\b(highest|maximum|max)\b", lower) else canonical_min
				if ref is None:
					claims_unverifiable += 1
					error_mapping += 1
				elif any(qca_is_close(number, ref, approximate) for number in numbers):
					claims_correct += 1
				else:
					claims_incorrect += 1
					error_maxmin += 1

		if re.search(r"\b(from|between|range|ranges|ranging|spans)\b", lower) and not skip_structural:
			claims_total += 1
			if len(numbers) < 2 or canonical_min is None or canonical_max is None:
				claims_unverifiable += 1
				error_mapping += 1
			else:
				claims_verifiable += 1
				n_min = min(numbers[0], numbers[1])
				n_max = max(numbers[0], numbers[1])
				ok_min = qca_is_close(n_min, float(canonical_min), approximate)
				ok_max = qca_is_close(n_max, float(canonical_max), approximate)
				if ok_min and ok_max:
					claims_correct += 1
				else:
					claims_incorrect += 1
					error_range += 1

		if re.search(r"\b(higher than|greater than|lower than|less than|more than|fewer than)\b", lower) and not skip_structural:
			claims_total += 1
			if len(numbers) < 2:
				claims_unverifiable += 1
				error_mapping += 1
			else:
				claims_verifiable += 1
				a, b = numbers[0], numbers[1]
				relation_ok = False
				if re.search(r"\b(higher than|greater than|more than)\b", lower):
					relation_ok = a > b
				elif re.search(r"\b(lower than|less than|fewer than)\b", lower):
					relation_ok = a < b
				in_canonical = any(qca_is_close(a, value, approximate) for value in values) and any(
					qca_is_close(b, value, approximate) for value in values
				)
				if relation_ok and in_canonical:
					claims_correct += 1
				else:
					claims_incorrect += 1
					error_comparison += 1

		trend_hit = None
		for trend_name, terms in QCA_TREND_PATTERNS.items():
			if any(term in lower for term in terms):
				trend_hit = trend_name
				break
		if trend_hit:
			claims_total += 1
			if canonical_trend in (None, "mixed"):
				claims_unverifiable += 1
				error_mapping += 1
			else:
				claims_verifiable += 1
				if trend_hit == canonical_trend:
					claims_correct += 1
				elif trend_hit == "increasing" and canonical_trend == "increase_then_decrease":
					claims_incorrect += 1
					error_trend += 1
				elif trend_hit == "decreasing" and canonical_trend == "decrease_then_increase":
					claims_incorrect += 1
					error_trend += 1
				else:
					claims_incorrect += 1
					error_trend += 1

		if numbers:
			has_structural_claim = bool(
				re.search(r"\b(highest|maximum|max|lowest|minimum|min|from|between|range|ranges|ranging|spans|higher than|greater than|lower than|less than|more than|fewer than)\b", lower)
			)
			if not has_structural_claim:
				for number in numbers:
					claims_total += 1
					claims_verifiable += 1
					if any(qca_is_close(number, value, approximate) for value in values):
						claims_correct += 1
					else:
						claims_incorrect += 1
						error_value += 1

	if claims_verifiable == 0:
		qca_score = None
		qca_status = "no_verifiable_claims"
	else:
		qca_score = round(claims_correct / claims_verifiable, 6)
		qca_status = "ok"

	return {
		"qca_score": qca_score,
		"qca_status": qca_status,
		"qca_claims_total": claims_total,
		"qca_claims_verifiable": claims_verifiable,
		"qca_claims_correct": claims_correct,
		"qca_claims_incorrect": claims_incorrect,
		"qca_claims_unverifiable": claims_unverifiable,
		"qca_error_value": error_value,
		"qca_error_maxmin": error_maxmin,
		"qca_error_comparison": error_comparison,
		"qca_error_range": error_range,
		"qca_error_trend": error_trend,
		"qca_error_mapping": error_mapping,
	}


def evaluate_dvr(description_text: str, canonical_case: Dict[str, Any] | None) -> Dict[str, Any]:
	if not canonical_case:
		return {
			"dvr_score": None,
			"dvr_status": "missing_canonical_case",
			"dvr_mentions_total": 0,
			"dvr_mentions_verifiable": 0,
			"dvr_mentions_valid": 0,
			"dvr_mentions_invalid": 0,
			"dvr_mentions_unverifiable": 0,
		}

	values = canonical_case.get("values") or []
	if not values:
		return {
			"dvr_score": None,
			"dvr_status": "missing_canonical_values",
			"dvr_mentions_total": 0,
			"dvr_mentions_verifiable": 0,
			"dvr_mentions_valid": 0,
			"dvr_mentions_invalid": 0,
			"dvr_mentions_unverifiable": 0,
		}

	mentions_total = 0
	mentions_verifiable = 0
	mentions_valid = 0
	mentions_invalid = 0
	mentions_unverifiable = 0

	for sentence in [segment.strip() for segment in SENTENCE_SPLIT_PATTERN.split(description_text) if segment.strip()]:
		raw_numbers = extract_numeric_values(sentence)
		numbers = filter_non_data_numbers(sentence, filter_value_numbers(raw_numbers, canonical_case))
		if not numbers:
			continue
		lower = sentence.lower()
		approximate = any(term in lower for term in QCA_APPROX_TERMS)
		for number in numbers:
			mentions_total += 1
			mentions_verifiable += 1
			if any(qca_is_close(number, value, approximate) for value in values):
				mentions_valid += 1
			else:
				mentions_invalid += 1

	if mentions_verifiable == 0:
		dvr_score = None
		dvr_status = "no_numeric_mentions"
	else:
		dvr_score = round(mentions_valid / mentions_verifiable, 6)
		dvr_status = "ok"

	return {
		"dvr_score": dvr_score,
		"dvr_status": dvr_status,
		"dvr_mentions_total": mentions_total,
		"dvr_mentions_verifiable": mentions_verifiable,
		"dvr_mentions_valid": mentions_valid,
		"dvr_mentions_invalid": mentions_invalid,
		"dvr_mentions_unverifiable": mentions_unverifiable,
	}


def count_number_words(text: str) -> int:
	normalized = text.lower()
	normalized = normalized.replace("—", " ").replace("–", " ")
	normalized = normalized.replace("-", " ")
	tokens = re.findall(r"[a-z]+", normalized)

	count = 0
	index = 0

	while index < len(tokens):
		token = tokens[index]

		if token in ORDINAL_WORDS:
			count += 1
			index += 1
			continue

		if token not in CARDINAL_NUMBER_WORDS:
			index += 1
			continue

		index += 1
		while index < len(tokens):
			current = tokens[index]
			if current in CARDINAL_NUMBER_WORDS:
				index += 1
				continue

			if current == "and" and index + 1 < len(tokens) and tokens[index + 1] in CARDINAL_NUMBER_WORDS:
				index += 1
				continue

			break

		count += 1

	return count


def words_per_number(word_count: int, number_count: int) -> float | None:
	if number_count == 0:
		return None
	return round(word_count / number_count, 6)


def build_description(parts: Iterable[Any]) -> str:
	clean_parts = [normalize_spaces(str(part)) for part in parts if part not in (None, "")]
	return " ".join(part for part in clean_parts if part)


def analyse_row(
	row: tuple[Any, ...],
	headers: tuple[Any, ...],
	canonical_cases: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
	row_map = {str(header).strip(): value for header, value in zip(headers, row) if header is not None}

	chart_label = str(row_map.get("Chart") or "")
	description_columns = [
		"Overview and main message",
		"Chart structure",
		"Relevant patterns, trends, and comparisons",
		"Essential key details",
	]
	description_text = build_description(row_map.get(column) for column in description_columns)

	word_count = len(tokenize_words(description_text))
	sentence_count = count_sentences(description_text)
	polysyllabic_word_count = count_polysyllabic_words(description_text)
	total_syllables = count_total_syllables(description_text)
	total_chars = count_characters(description_text)
	numeric_token_count = count_numeric_tokens(description_text)
	number_word_count = count_number_words(description_text)
	total_number_mentions = numeric_token_count + number_word_count
	smog = smog_index(polysyllabic_word_count, sentence_count)
	gunning_fog = gunning_fog_index(word_count, sentence_count, polysyllabic_word_count)
	fk_grade = flesch_kincaid_grade(word_count, sentence_count, total_syllables)
	cl_index = coleman_liau_index(total_chars, sentence_count, word_count)
	dict_confidence = count_dictionary_confidence(description_text)
	
	# Content Structure Detection
	chart_type = has_chart_type(description_text)
	title = has_title(description_text)
	axis_labels = has_axis_labels(description_text)
	categories = has_categories(description_text)
	values = has_values(description_text)
	scale_info = has_scale_info(description_text)
	ics_score = round(sum([chart_type, title, axis_labels, categories, values, scale_info]) / 6 * 100, 2)
	case_id = extract_case_id(chart_label)
	qca = evaluate_qca(description_text, canonical_cases.get(case_id))
	dvr = evaluate_dvr(description_text, canonical_cases.get(case_id))
	if dvr["dvr_mentions_verifiable"] == 0:
		hdr_score = None
		hdr_status = dvr["dvr_status"]
	else:
		hdr_score = round(dvr["dvr_mentions_invalid"] / dvr["dvr_mentions_verifiable"], 6)
		hdr_status = "ok"

	return {
		"chart": chart_label,
		"id": case_id,
		"dvr_score": dvr["dvr_score"],
		"hdr_score": hdr_score,
		"qca_fair_score": qca["qca_score"],
		"ics_score": ics_score,
		"nd_score": round((total_number_mentions / word_count) * 100, 6) if word_count else None,
		"smog_index": smog,
		"gunning_fog_index": gunning_fog,
		"flesch_kincaid_grade": fk_grade,
		"coleman_liau_index": cl_index,
		"dictionary_confidence": dict_confidence,
		"word_count": word_count,
		"sentence_count": sentence_count,
		"polysyllabic_word_count": polysyllabic_word_count,
		"total_syllables": total_syllables,
		"total_characters": total_chars,
		"numeric_token_count": numeric_token_count,
		"number_word_count": number_word_count,
		"total_number_mentions": total_number_mentions,
		"words_per_number": words_per_number(word_count, total_number_mentions),
		"has_chart_type": chart_type,
		"has_title": title,
		"has_axis_labels": axis_labels,
		"has_categories": categories,
		"has_values": values,
		"has_scale_info": scale_info,
		"qca_fair_status": qca["qca_status"],
		"qca_fair_claims_total": qca["qca_claims_total"],
		"qca_fair_claims_verifiable": qca["qca_claims_verifiable"],
		"qca_fair_claims_correct": qca["qca_claims_correct"],
		"qca_fair_claims_incorrect": qca["qca_claims_incorrect"],
		"qca_fair_claims_unverifiable": qca["qca_claims_unverifiable"],
		"qca_fair_error_value": qca["qca_error_value"],
		"qca_fair_error_maxmin": qca["qca_error_maxmin"],
		"qca_fair_error_comparison": qca["qca_error_comparison"],
		"qca_fair_error_range": qca["qca_error_range"],
		"qca_fair_error_trend": qca["qca_error_trend"],
		"qca_fair_error_mapping": qca["qca_error_mapping"],
		"dvr_status": dvr["dvr_status"],
		"dvr_mentions_total": dvr["dvr_mentions_total"],
		"dvr_mentions_verifiable": dvr["dvr_mentions_verifiable"],
		"dvr_mentions_valid": dvr["dvr_mentions_valid"],
		"dvr_mentions_invalid": dvr["dvr_mentions_invalid"],
		"dvr_mentions_unverifiable": dvr["dvr_mentions_unverifiable"],
		"hdr_status": hdr_status,
		"hdr_mentions_total": dvr["dvr_mentions_total"],
		"hdr_mentions_verifiable": dvr["dvr_mentions_verifiable"],
		"hdr_mentions_hallucinated": dvr["dvr_mentions_invalid"],
		"hdr_mentions_non_hallucinated": dvr["dvr_mentions_valid"],
		"hdr_mentions_unverifiable": dvr["dvr_mentions_unverifiable"],
		"description_text": description_text,
	}


def mean(values: List[float]) -> float | None:
	if not values:
		return None
	return round(sum(values) / len(values), 6)


def median(values: List[float]) -> float | None:
	if not values:
		return None
	sorted_values = sorted(values)
	mid = len(sorted_values) // 2
	if len(sorted_values) % 2 == 1:
		return round(sorted_values[mid], 6)
	return round((sorted_values[mid - 1] + sorted_values[mid]) / 2, 6)


def stdev_sample(values: List[float]) -> float | None:
	if len(values) < 2:
		return 0.0 if values else None
	avg = sum(values) / len(values)
	variance = sum((value - avg) ** 2 for value in values) / (len(values) - 1)
	return round(math.sqrt(variance), 6)


def classify_high_better(value: float | None, cutoffs: List[float], labels: List[str]) -> str:
	if value is None:
		return "not_available"
	if value < cutoffs[0]:
		return labels[0]
	if value < cutoffs[1]:
		return labels[1]
	if value < cutoffs[2]:
		return labels[2]
	return labels[3]


def classify_low_better(value: float | None, cutoffs: List[float], labels: List[str]) -> str:
	if value is None:
		return "not_available"
	if value <= cutoffs[0]:
		return labels[0]
	if value <= cutoffs[1]:
		return labels[1]
	if value <= cutoffs[2]:
		return labels[2]
	return labels[3]


def with_metric_summary(metric_dict: Dict[str, Any], summary: Dict[str, Any]) -> Dict[str, Any]:
	return {
		"summary": summary,
		**metric_dict,
	}


def summarise_metrics(rows: List[Dict[str, Any]], input_file: str | None = None) -> Dict[str, Any]:
	word_counts = [float(row["word_count"]) for row in rows]
	sentence_counts = [float(row["sentence_count"]) for row in rows]
	polysyllabic_counts = [float(row["polysyllabic_word_count"]) for row in rows]
	smog_values = [float(row["smog_index"]) for row in rows if row["smog_index"] is not None]
	gunning_values = [float(row["gunning_fog_index"]) for row in rows if row["gunning_fog_index"] is not None]
	fk_values = [float(row["flesch_kincaid_grade"]) for row in rows if row["flesch_kincaid_grade"] is not None]
	cl_values = [float(row["coleman_liau_index"]) for row in rows if row["coleman_liau_index"] is not None]
	dict_confidence_values = [float(row["dictionary_confidence"]) for row in rows if row["dictionary_confidence"] is not None]
	total_numbers = [float(row["total_number_mentions"]) for row in rows]
	numeric_tokens = [float(row["numeric_token_count"]) for row in rows]
	number_words = [float(row["number_word_count"]) for row in rows]
	qca_fair_scores = [float(row["qca_fair_score"]) for row in rows if row["qca_fair_score"] is not None]
	qca_fair_verifiable_total = sum(int(row["qca_fair_claims_verifiable"]) for row in rows)
	qca_fair_total_claims = sum(int(row["qca_fair_claims_total"]) for row in rows)
	qca_fair_correct_total = sum(int(row["qca_fair_claims_correct"]) for row in rows)
	qca_fair_incorrect_total = sum(int(row["qca_fair_claims_incorrect"]) for row in rows)
	qca_fair_unverifiable_total = sum(int(row["qca_fair_claims_unverifiable"]) for row in rows)
	dvr_scores = [float(row["dvr_score"]) for row in rows if row["dvr_score"] is not None]
	dvr_mentions_total = sum(int(row["dvr_mentions_total"]) for row in rows)
	dvr_mentions_verifiable = sum(int(row["dvr_mentions_verifiable"]) for row in rows)
	dvr_mentions_valid = sum(int(row["dvr_mentions_valid"]) for row in rows)
	dvr_mentions_invalid = sum(int(row["dvr_mentions_invalid"]) for row in rows)
	dvr_mentions_unverifiable = sum(int(row["dvr_mentions_unverifiable"]) for row in rows)
	hdr_scores = [float(row["hdr_score"]) for row in rows if row["hdr_score"] is not None]
	hdr_mentions_total = sum(int(row["hdr_mentions_total"]) for row in rows)
	hdr_mentions_verifiable = sum(int(row["hdr_mentions_verifiable"]) for row in rows)
	hdr_mentions_hallucinated = sum(int(row["hdr_mentions_hallucinated"]) for row in rows)
	hdr_mentions_non_hallucinated = sum(int(row["hdr_mentions_non_hallucinated"]) for row in rows)
	hdr_mentions_unverifiable = sum(int(row["hdr_mentions_unverifiable"]) for row in rows)

	highest_ratio = max(rows, key=lambda row: row["words_per_number"] if row["words_per_number"] is not None else float("inf"))
	highest_density = max(rows, key=lambda row: row["nd_score"] if row["nd_score"] is not None else -1)
	longest_description = max(rows, key=lambda row: row["word_count"])
	most_numbers = max(rows, key=lambda row: row["total_number_mentions"])
	highest_smog = max(rows, key=lambda row: row["smog_index"] if row["smog_index"] is not None else -1)
	lowest_smog = min(rows, key=lambda row: row["smog_index"] if row["smog_index"] is not None else float("inf"))

	result = {
		"cases_analysed": len(rows),
		"word_count": {
			"mean": mean(word_counts),
			"median": median(word_counts),
			"min": min(word_counts) if word_counts else None,
			"max": max(word_counts) if word_counts else None,
			"sample_stdev": stdev_sample(word_counts),
		},
		"number_mentions": {
			"mean": mean(total_numbers),
			"median": median(total_numbers),
			"min": min(total_numbers) if total_numbers else None,
			"max": max(total_numbers) if total_numbers else None,
			"sample_stdev": stdev_sample(total_numbers),
			"mean_numeric_tokens": mean(numeric_tokens),
			"mean_number_words": mean(number_words),
		},
		"smog": {
			"mean": mean(smog_values),
			"median": median(smog_values),
			"min": min(smog_values) if smog_values else None,
			"max": max(smog_values) if smog_values else None,
			"sample_stdev": stdev_sample(smog_values),
			"mean_sentence_count": mean(sentence_counts),
			"mean_polysyllabic_word_count": mean(polysyllabic_counts),
			"mean_dictionary_confidence": mean(dict_confidence_values),
		},
		"gunning_fog": {
			"mean": mean(gunning_values),
			"median": median(gunning_values),
			"min": min(gunning_values) if gunning_values else None,
			"max": max(gunning_values) if gunning_values else None,
			"sample_stdev": stdev_sample(gunning_values),
		},
		"flesch_kincaid_grade": {
			"mean": mean(fk_values),
			"median": median(fk_values),
			"min": min(fk_values) if fk_values else None,
			"max": max(fk_values) if fk_values else None,
			"sample_stdev": stdev_sample(fk_values),
		},
		"coleman_liau": {
			"mean": mean(cl_values),
			"median": median(cl_values),
			"min": min(cl_values) if cl_values else None,
			"max": max(cl_values) if cl_values else None,
			"sample_stdev": stdev_sample(cl_values),
		},
		"content_structure": {
			"has_chart_type_pct": round(sum(1 for row in rows if row["has_chart_type"]) / len(rows) * 100, 2) if rows else None,
			"has_title_pct": round(sum(1 for row in rows if row["has_title"]) / len(rows) * 100, 2) if rows else None,
			"has_axis_labels_pct": round(sum(1 for row in rows if row["has_axis_labels"]) / len(rows) * 100, 2) if rows else None,
			"has_categories_pct": round(sum(1 for row in rows if row["has_categories"]) / len(rows) * 100, 2) if rows else None,
			"has_values_pct": round(sum(1 for row in rows if row["has_values"]) / len(rows) * 100, 2) if rows else None,
			"has_scale_info_pct": round(sum(1 for row in rows if row["has_scale_info"]) / len(rows) * 100, 2) if rows else None,
			"ics_score": {
				"mean": mean([float(row["ics_score"]) for row in rows]),
				"median": median([float(row["ics_score"]) for row in rows]),
				"min": min([row["ics_score"] for row in rows]) if rows else None,
				"max": max([row["ics_score"] for row in rows]) if rows else None,
			},
		},
		"qca_fair": {
			"mean": mean(qca_fair_scores),
			"median": median(qca_fair_scores),
			"min": min(qca_fair_scores) if qca_fair_scores else None,
			"max": max(qca_fair_scores) if qca_fair_scores else None,
			"sample_stdev": stdev_sample(qca_fair_scores),
			"claims_total": qca_fair_total_claims,
			"claims_verifiable": qca_fair_verifiable_total,
			"claims_correct": qca_fair_correct_total,
			"claims_incorrect": qca_fair_incorrect_total,
			"claims_unverifiable": qca_fair_unverifiable_total,
			"verifiable_rate": round((qca_fair_verifiable_total / qca_fair_total_claims) * 100, 2) if qca_fair_total_claims else None,
			"correct_rate_on_verifiable": round((qca_fair_correct_total / qca_fair_verifiable_total) * 100, 2) if qca_fair_verifiable_total else None,
			"error_breakdown": {
				"value_error": sum(int(row["qca_fair_error_value"]) for row in rows),
				"maxmin_error": sum(int(row["qca_fair_error_maxmin"]) for row in rows),
				"comparison_error": sum(int(row["qca_fair_error_comparison"]) for row in rows),
				"range_error": sum(int(row["qca_fair_error_range"]) for row in rows),
				"trend_error": sum(int(row["qca_fair_error_trend"]) for row in rows),
				"mapping_error": sum(int(row["qca_fair_error_mapping"]) for row in rows),
			},
		},
		"dvr": {
			"mean": mean(dvr_scores),
			"median": median(dvr_scores),
			"min": min(dvr_scores) if dvr_scores else None,
			"max": max(dvr_scores) if dvr_scores else None,
			"sample_stdev": stdev_sample(dvr_scores),
			"mentions_total": dvr_mentions_total,
			"mentions_verifiable": dvr_mentions_verifiable,
			"mentions_valid": dvr_mentions_valid,
			"mentions_invalid": dvr_mentions_invalid,
			"mentions_unverifiable": dvr_mentions_unverifiable,
			"valid_rate_on_verifiable": round((dvr_mentions_valid / dvr_mentions_verifiable) * 100, 2) if dvr_mentions_verifiable else None,
		},
		"hdr": {
			"mean": mean(hdr_scores),
			"median": median(hdr_scores),
			"min": min(hdr_scores) if hdr_scores else None,
			"max": max(hdr_scores) if hdr_scores else None,
			"sample_stdev": stdev_sample(hdr_scores),
			"mentions_total": hdr_mentions_total,
			"mentions_verifiable": hdr_mentions_verifiable,
			"mentions_hallucinated": hdr_mentions_hallucinated,
			"mentions_non_hallucinated": hdr_mentions_non_hallucinated,
			"hallucinated_rate_on_verifiable": round((hdr_mentions_hallucinated / hdr_mentions_verifiable) * 100, 2) if hdr_mentions_verifiable else None,
		},
		"extremes": {
			"longest_description": {
				"id": longest_description["id"],
				"chart": longest_description["chart"],
				"word_count": longest_description["word_count"],
			},
			"most_number_mentions": {
				"id": most_numbers["id"],
				"chart": most_numbers["chart"],
				"total_number_mentions": most_numbers["total_number_mentions"],
			},
			"highest_words_per_number": {
				"id": highest_ratio["id"],
				"chart": highest_ratio["chart"],
				"words_per_number": highest_ratio["words_per_number"],
			},
			"highest_number_density": {
				"id": highest_density["id"],
				"chart": highest_density["chart"],
				"nd_score": highest_density["nd_score"],
			},
			"highest_smog": {
				"id": highest_smog["id"],
				"chart": highest_smog["chart"],
				"smog_index": highest_smog["smog_index"],
			},
			"lowest_smog": {
				"id": lowest_smog["id"],
				"chart": lowest_smog["chart"],
				"smog_index": lowest_smog["smog_index"],
			},
		},
	}

	# Add metric-level reference values and interpretation summaries.
	result["qca_fair"] = with_metric_summary(
		result["qca_fair"],
		{
			"metric_label": "QCA_fair",
			"direction": "higher_is_better",
			"reference_ranges": {
				"low": "< 0.40",
				"moderate": "0.40-0.60",
				"good": "0.60-0.80",
				"excellent": ">= 0.80",
			},
			"interpretation": classify_high_better(
				result["qca_fair"].get("mean"),
				[0.40, 0.60, 0.80],
				["low", "moderate", "good", "excellent"],
			),
		},
	)

	result["dvr"] = with_metric_summary(
		result["dvr"],
		{
			"metric_label": "DVR",
			"direction": "higher_is_better",
			"reference_ranges": {
				"low": "< 0.40",
				"moderate": "0.40-0.60",
				"good": "0.60-0.80",
				"excellent": ">= 0.80",
			},
			"interpretation": classify_high_better(
				result["dvr"].get("mean"),
				[0.40, 0.60, 0.80],
				["low", "moderate", "good", "excellent"],
			),
		},
	)

	result["hdr"] = with_metric_summary(
		result["hdr"],
		{
			"metric_label": "HDR",
			"direction": "lower_is_better",
			"reference_ranges": {
				"excellent": "<= 0.20",
				"good": "0.20-0.35",
				"moderate": "0.35-0.50",
				"high_risk": "> 0.50",
			},
			"interpretation": classify_low_better(
				result["hdr"].get("mean"),
				[0.20, 0.35, 0.50],
				["excellent", "good", "moderate", "high_risk"],
			),
		},
	)

	result["content_structure"]["ics_score"] = with_metric_summary(
		result["content_structure"]["ics_score"],
		{
			"metric_label": "ICS",
			"direction": "higher_is_better",
			"reference_ranges": {
				"low": "< 65",
				"moderate": "65-80",
				"good": "80-90",
				"excellent": ">= 90",
			},
			"interpretation": classify_high_better(
				result["content_structure"]["ics_score"].get("mean"),
				[65.0, 80.0, 90.0],
				["low", "moderate", "good", "excellent"],
			),
		},
	)

	result["smog"] = with_metric_summary(
		result["smog"],
		{
			"metric_label": "SMOG",
			"direction": "lower_is_easier",
			"reference_ranges": {
				"easy": "< 7",
				"moderate": "7-9",
				"difficult": "9-11",
				"very_difficult": ">= 11",
			},
			"interpretation": classify_low_better(
				result["smog"].get("mean"),
				[7.0, 9.0, 11.0],
				["easy", "moderate", "difficult", "very_difficult"],
			),
		},
	)

	result["gunning_fog"] = with_metric_summary(
		result["gunning_fog"],
		{
			"metric_label": "Gunning Fog",
			"direction": "lower_is_easier",
			"reference_ranges": {
				"easy": "< 10",
				"moderate": "10-12",
				"difficult": "12-15",
				"very_difficult": ">= 15",
			},
			"interpretation": classify_low_better(
				result["gunning_fog"].get("mean"),
				[10.0, 12.0, 15.0],
				["easy", "moderate", "difficult", "very_difficult"],
			),
		},
	)

	result["flesch_kincaid_grade"] = with_metric_summary(
		result["flesch_kincaid_grade"],
		{
			"metric_label": "Flesch-Kincaid Grade",
			"direction": "lower_is_easier",
			"reference_ranges": {
				"easy": "< 8",
				"moderate": "8-10",
				"difficult": "10-12",
				"very_difficult": ">= 12",
			},
			"interpretation": classify_low_better(
				result["flesch_kincaid_grade"].get("mean"),
				[8.0, 10.0, 12.0],
				["easy", "moderate", "difficult", "very_difficult"],
			),
		},
	)

	result["coleman_liau"] = with_metric_summary(
		result["coleman_liau"],
		{
			"metric_label": "Coleman-Liau",
			"direction": "lower_is_easier",
			"reference_ranges": {
				"easy": "< 10",
				"moderate": "10-13",
				"difficult": "13-16",
				"very_difficult": ">= 16",
			},
			"interpretation": classify_low_better(
				result["coleman_liau"].get("mean"),
				[10.0, 13.0, 16.0],
				["easy", "moderate", "difficult", "very_difficult"],
			),
		},
	)
	
	return result


def load_metrics(canonical_cases: Dict[str, Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
	"""Load metrics for all configured providers."""
	results = {}
	
	for provider_name, provider_path in PROVIDERS.items():
		if not provider_path.exists():
			raise FileNotFoundError(f"No s'ha trobat el fitxer per a {provider_name}: {provider_path}")

		workbook = load_workbook(provider_path, read_only=True, data_only=True)
		worksheet = workbook[workbook.sheetnames[0]]

		rows_iter = worksheet.iter_rows(values_only=True)
		headers = next(rows_iter)
		provider_results: List[Dict[str, Any]] = []

		for row in rows_iter:
			if not row or all(value in (None, "") for value in row):
				continue
			provider_results.append(analyse_row(row, headers, canonical_cases))

		workbook.close()
		results[provider_name] = provider_results
	
	return results


def write_csv(rows: List[Dict[str, Any]]) -> None:
	OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
	headers = [
		"id",
		"chart",
		"word_count",
		"sentence_count",
		"polysyllabic_word_count",
		"total_syllables",
		"total_characters",
		"smog_index",
		"gunning_fog_index",
		"flesch_kincaid_grade",
		"coleman_liau_index",
		"dictionary_confidence",
		"numeric_token_count",
		"number_word_count",
		"total_number_mentions",
		"words_per_number",
		"nd_score",
	]
	with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as handle:
		writer = csv.DictWriter(handle, fieldnames=headers)
		writer.writeheader()
		for row in rows:
			writer.writerow({header: row.get(header) for header in headers})


def build_unified_cases(all_metrics: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
	"""Combine metrics from all providers into unified case records."""
	# Group by case ID
	cases_by_id = {}
	
	for provider_name, cases in all_metrics.items():
		for case in cases:
			case_id = case["id"]
			if case_id not in cases_by_id:
				cases_by_id[case_id] = {"id": case_id, "chart": case["chart"]}
			
			# Add provider-prefixed metrics
			for key, value in case.items():
				if key not in ("id", "chart", "description_text"):
					prefixed_key = f"{provider_name}_{key}"
					cases_by_id[case_id][prefixed_key] = value
	
	return sorted(cases_by_id.values(), key=lambda x: x["id"])


def write_csv_unified(unified_cases: List[Dict[str, Any]]) -> None:
	"""Write unified CSV with metrics from all providers."""
	OUTPUT_CSV_UNIFIED.parent.mkdir(parents=True, exist_ok=True)
	
	# Determine all possible headers
	all_keys = set()
	for case in unified_cases:
		all_keys.update(case.keys())
	
	# Score metrics first (in desired order), then remaining data fields alphabetically
	_SCORE_ORDER = [
		"dvr_score", "hdr_score", "qca_fair_score", "ics_score", "nd_score",
		"smog_index", "gunning_fog_index", "flesch_kincaid_grade", "coleman_liau_index",
		"dictionary_confidence",
	]
	headers = ["id", "chart"]
	remaining = set(k for k in all_keys if k not in ("id", "chart"))
	for metric in _SCORE_ORDER:
		for p in PROVIDERS:
			key = f"{p}_{metric}"
			if key in remaining:
				headers.append(key)
				remaining.discard(key)
	headers.extend(sorted(remaining))
	
	with OUTPUT_CSV_UNIFIED.open("w", encoding="utf-8", newline="") as handle:
		writer = csv.DictWriter(handle, fieldnames=headers)
		writer.writeheader()
		for case in unified_cases:
			writer.writerow({header: case.get(header) for header in headers})


def write_json_unified(unified_cases: List[Dict[str, Any]]) -> None:
	"""Write unified JSON with metrics from all providers."""
	OUTPUT_JSON_UNIFIED.parent.mkdir(parents=True, exist_ok=True)
	payload = {
		"providers": list(PROVIDERS.keys()),
		"cases": unified_cases,
	}
	with OUTPUT_JSON_UNIFIED.open("w", encoding="utf-8") as handle:
		json.dump(payload, handle, ensure_ascii=False, indent=2)


def summarise_per_provider(all_metrics: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Dict[str, Any]]:
	"""Generate aggregated metrics per provider."""
	summaries = {}
	for provider_name, rows in all_metrics.items():
		summaries[provider_name] = summarise_metrics(rows)
	return summaries


def write_csv_aggregated(provider_summaries: Dict[str, Dict[str, Any]]) -> None:
	"""Write aggregated metrics as CSV (one row per provider per metric)."""
	OUTPUT_CSV_AGGREGATED.parent.mkdir(parents=True, exist_ok=True)
	
	rows_to_write = []
	
	for provider_name, summary in provider_summaries.items():
		# Word count metrics
		rows_to_write.append({
			"provider": provider_name,
			"metric": "word_count",
			"mean": summary["word_count"]["mean"],
			"median": summary["word_count"]["median"],
			"min": summary["word_count"]["min"],
			"max": summary["word_count"]["max"],
			"stdev": summary["word_count"]["sample_stdev"],
		})
		
		# SMOG
		rows_to_write.append({
			"provider": provider_name,
			"metric": "smog_index",
			"mean": summary["smog"]["mean"],
			"median": summary["smog"]["median"],
			"min": summary["smog"]["min"],
			"max": summary["smog"]["max"],
			"stdev": summary["smog"]["sample_stdev"],
		})
		
		# Gunning Fog
		rows_to_write.append({
			"provider": provider_name,
			"metric": "gunning_fog_index",
			"mean": summary["gunning_fog"]["mean"],
			"median": summary["gunning_fog"]["median"],
			"min": summary["gunning_fog"]["min"],
			"max": summary["gunning_fog"]["max"],
			"stdev": summary["gunning_fog"]["sample_stdev"],
		})
		
		# Flesch-Kincaid
		rows_to_write.append({
			"provider": provider_name,
			"metric": "flesch_kincaid_grade",
			"mean": summary["flesch_kincaid_grade"]["mean"],
			"median": summary["flesch_kincaid_grade"]["median"],
			"min": summary["flesch_kincaid_grade"]["min"],
			"max": summary["flesch_kincaid_grade"]["max"],
			"stdev": summary["flesch_kincaid_grade"]["sample_stdev"],
		})
		
		# Coleman-Liau
		rows_to_write.append({
			"provider": provider_name,
			"metric": "coleman_liau_index",
			"mean": summary["coleman_liau"]["mean"],
			"median": summary["coleman_liau"]["median"],
			"min": summary["coleman_liau"]["min"],
			"max": summary["coleman_liau"]["max"],
			"stdev": summary["coleman_liau"]["sample_stdev"],
		})
		
		# Number mentions
		rows_to_write.append({
			"provider": provider_name,
			"metric": "total_number_mentions",
			"mean": summary["number_mentions"]["mean"],
			"median": summary["number_mentions"]["median"],
			"min": summary["number_mentions"]["min"],
			"max": summary["number_mentions"]["max"],
			"stdev": summary["number_mentions"]["sample_stdev"],
		})

		# QCA
		rows_to_write.append({
			"provider": provider_name,
			"metric": "qca_fair_score",
			"mean": summary["qca_fair"]["mean"],
			"median": summary["qca_fair"]["median"],
			"min": summary["qca_fair"]["min"],
			"max": summary["qca_fair"]["max"],
			"stdev": summary["qca_fair"]["sample_stdev"],
		})

		# DVR
		rows_to_write.append({
			"provider": provider_name,
			"metric": "dvr_score",
			"mean": summary["dvr"]["mean"],
			"median": summary["dvr"]["median"],
			"min": summary["dvr"]["min"],
			"max": summary["dvr"]["max"],
			"stdev": summary["dvr"]["sample_stdev"],
		})

		# HDR
		rows_to_write.append({
			"provider": provider_name,
			"metric": "hdr_score",
			"mean": summary["hdr"]["mean"],
			"median": summary["hdr"]["median"],
			"min": summary["hdr"]["min"],
			"max": summary["hdr"]["max"],
			"stdev": summary["hdr"]["sample_stdev"],
		})
	
	headers = ["provider", "metric", "mean", "median", "min", "max", "stdev"]
	with OUTPUT_CSV_AGGREGATED.open("w", encoding="utf-8", newline="") as handle:
		writer = csv.DictWriter(handle, fieldnames=headers)
		writer.writeheader()
		writer.writerows(rows_to_write)


def write_json_aggregated(provider_summaries: Dict[str, Dict[str, Any]]) -> None:
	"""Write aggregated metrics as JSON per provider."""
	OUTPUT_JSON_AGGREGATED.parent.mkdir(parents=True, exist_ok=True)
	payload = {provider_name: summary for provider_name, summary in provider_summaries.items()}
	with OUTPUT_JSON_AGGREGATED.open("w", encoding="utf-8") as handle:
		json.dump(payload, handle, ensure_ascii=False, indent=2)


# Keep backwards-compatible functions
def write_csv(rows: List[Dict[str, Any]]) -> None:
	OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
	headers = [
		"id",
		"chart",
		"word_count",
		"sentence_count",
		"polysyllabic_word_count",
		"total_syllables",
		"total_characters",
		"smog_index",
		"gunning_fog_index",
		"flesch_kincaid_grade",
		"coleman_liau_index",
		"dictionary_confidence",
		"numeric_token_count",
		"number_word_count",
		"total_number_mentions",
		"words_per_number",
		"nd_score",
	]
	with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as handle:
		writer = csv.DictWriter(handle, fieldnames=headers)
		writer.writeheader()
		for row in rows:
			writer.writerow({header: row.get(header) for header in headers})


def write_json(rows: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
	OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
	payload = {
		"summary": summary,
		"cases": rows,
	}
	with OUTPUT_JSON.open("w", encoding="utf-8") as handle:
		json.dump(payload, handle, ensure_ascii=False, indent=2)


def main() -> None:
	canonical_cases = load_canonical_cases()
	# Load metrics for all providers
	all_metrics = load_metrics(canonical_cases)
	
	# Generate unified case data
	unified_cases = build_unified_cases(all_metrics)
	
	# Generate provider summaries
	provider_summaries = summarise_per_provider(all_metrics)
	
	# Write unified outputs
	write_csv_unified(unified_cases)
	write_json_unified(unified_cases)
	
	# Write aggregated outputs
	write_csv_aggregated(provider_summaries)
	write_json_aggregated(provider_summaries)
	
	# Print summary information
	print("=" * 80)
	print("RESUM DE MÈTRIQUES MULTI-PROVEÏDOR")
	print("=" * 80)
	for provider_name, summary in provider_summaries.items():
		print(f"\n{provider_name.upper()}:")
		print(f"  Casos analitzats: {summary['cases_analysed']}")
		print(f"  Mitjana paraules: {summary['word_count']['mean']}")
		print(f"  Mitjana mencions numèriques: {summary['number_mentions']['mean']}")
		print(f"  Mitjana SMOG: {summary['smog']['mean']}")
		print(f"  Mitjana QCA_fair: {summary['qca_fair']['mean']}")
		print(f"  Mitjana DVR: {summary['dvr']['mean']}")
		print(f"  Mitjana HDR: {summary['hdr']['mean']}")
	
	print("\n" + "=" * 80)
	print("FITXERS GENERATS:")
	print("=" * 80)
	print(f"CSV unificat: {OUTPUT_CSV_UNIFIED}")
	print(f"JSON unificat: {OUTPUT_JSON_UNIFIED}")
	print(f"CSV agregat: {OUTPUT_CSV_AGGREGATED}")
	print(f"JSON agregat: {OUTPUT_JSON_AGGREGATED}")


if __name__ == "__main__":
	main()
